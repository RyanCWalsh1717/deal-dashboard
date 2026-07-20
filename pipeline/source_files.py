"""Auto-classifies uploaded source files (by content, not filename/extension
alone) and saves them into period-versioned folders (`source_files/<YYYY-MM>/`)
so a prior month's snapshot stays browsable instead of being overwritten by
the next upload.

Classification reuses the SAME parsers that will ultimately read each file —
every parser already returns an empty/None result on a shape mismatch rather
than raising, so trying them in priority order IS the classifier. There is no
separate, parallel sniffing implementation to keep in sync with the real
parsers.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import List, Optional

import openpyxl

from pipeline.parsers._shared import parse_as_of_date
from pipeline.parsers.cash_accounts import parse_loan_statement, parse_trial_balance_cash_accounts
from pipeline.parsers.rent_roll import parse_rent_roll
from pipeline.parsers.waterfall import parse_distribution_waterfall
from pipeline.property_config import PropertyConfig

_PERIOD_RE = re.compile(r"^\d{4}-\d{2}$")

_CANONICAL_NAMES = {
    "distribution_workbook": "distribution_workbook.xlsx",
    "trial_balance": "trial_balance.xlsx",
    "rent_roll": "rent_roll.xlsx",
}


@dataclass
class ClassifiedUpload:
    file_type: str  # "distribution_workbook" | "trial_balance" | "rent_roll" | "loan_statement" | "unknown"
    period: Optional[str] = None  # "YYYY-MM"
    tranche_name: Optional[str] = None  # loan_statement only
    error: Optional[str] = None


def _to_period(d: Optional[date]) -> Optional[str]:
    return f"{d.year:04d}-{d.month:02d}" if d else None


def _slugify(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", name.strip().lower()).strip("_") or "tranche"


def _workbook_as_of(cfg: PropertyConfig, wb) -> Optional[date]:
    """Mirrors the as-of derivation `distribution_workbook.parse_workbook()`
    does internally (top-level waterfall tier's "As of <Month> <Year>" label)
    — duplicated here rather than imported, since parse_workbook() only uses
    that date locally and doesn't expose it on DistributionWorkbookResult."""
    tier = cfg.top_level_tier()
    sheet_name = cfg.sheet_map.get("waterfall_property")
    if not tier or not sheet_name or sheet_name not in wb.sheetnames:
        return None
    tier_result = parse_distribution_waterfall(wb[sheet_name], tier.tier_id, tier.distributing_entity)
    return parse_as_of_date(tier_result.as_of_label)


def classify_upload(data: bytes, filename: str, cfg: PropertyConfig) -> ClassifiedUpload:
    ext = Path(filename).suffix.lower()

    if ext == ".pdf":
        stmt = parse_loan_statement(io.BytesIO(data))
        if stmt is None:
            return ClassifiedUpload(
                file_type="unknown", error=f'Couldn\'t recognize "{filename}" as a loan statement.'
            )
        return ClassifiedUpload(
            file_type="loan_statement",
            period=_to_period(stmt.as_of),
            tranche_name=stmt.tranche_name,
            error=None
            if stmt.as_of
            else f'"{filename}" parsed, but no "AS OF" date was found — can\'t file it under a period.',
        )

    if ext != ".xlsx":
        return ClassifiedUpload(file_type="unknown", error=f'"{filename}" isn\'t a .xlsx or .pdf file.')

    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)

    if any(name in wb.sheetnames for name in cfg.sheet_map.values()):
        as_of = _workbook_as_of(cfg, wb)
        return ClassifiedUpload(
            file_type="distribution_workbook",
            period=_to_period(as_of),
            error=None
            if as_of
            else f'"{filename}" looks like the distribution workbook, but no "As of <Month> <Year>" '
            "label was found on the top-level waterfall tab.",
        )

    rent_roll_ws = wb["Report1"] if "Report1" in wb.sheetnames else wb.worksheets[0]
    rent_roll_result = parse_rent_roll(rent_roll_ws, cfg.property_code)
    if rent_roll_result.lines:
        return ClassifiedUpload(
            file_type="rent_roll",
            period=_to_period(rent_roll_result.as_of),
            error=None
            if rent_roll_result.as_of
            else f'"{filename}" looks like a rent roll, but no "As of Date:" label was found.',
        )

    tb_accounts = parse_trial_balance_cash_accounts(io.BytesIO(data), cfg.yardi_codes)
    if tb_accounts:
        as_of = tb_accounts[0].as_of
        return ClassifiedUpload(
            file_type="trial_balance",
            period=_to_period(as_of),
            error=None
            if as_of
            else f'"{filename}" looks like a trial balance, but no "Period = <Month Year>" label was found.',
        )

    return ClassifiedUpload(
        file_type="unknown", error=f'Couldn\'t recognize "{filename}" as any known source file type.'
    )


def save_classified_upload(
    classified: ClassifiedUpload, data: bytes, cfg: PropertyConfig, data_dir: str = "data"
) -> Path:
    if classified.file_type == "unknown" or not classified.period:
        raise ValueError(classified.error or "Cannot save an unclassified upload.")

    period_dir = Path(data_dir) / cfg.property_code / "source_files" / classified.period
    if classified.file_type == "loan_statement":
        target = period_dir / "loan_statements" / f"{_slugify(classified.tranche_name or 'tranche')}.pdf"
    else:
        target = period_dir / _CANONICAL_NAMES[classified.file_type]

    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_bytes(data)
    return target


def list_periods(cfg: PropertyConfig, data_dir: str = "data") -> List[str]:
    base = Path(data_dir) / cfg.property_code / "source_files"
    if not base.exists():
        return []
    periods = [p.name for p in base.iterdir() if p.is_dir() and _PERIOD_RE.match(p.name)]
    return sorted(periods, reverse=True)


def resolve_period_file(
    cfg: PropertyConfig, period: str, relative_name: str, data_dir: str = "data"
) -> Optional[Path]:
    """Carry-forward lookup: `relative_name` in `period`'s folder, else the
    nearest earlier period that has it, else None — so a file type that
    wasn't re-uploaded every month still shows the last known version instead
    of going blank."""
    base = Path(data_dir) / cfg.property_code / "source_files"
    for candidate_period in list_periods(cfg, data_dir):
        if candidate_period > period:
            continue
        candidate = base / candidate_period / relative_name
        if candidate.exists():
            return candidate
    return None


def resolve_period_loan_statements(cfg: PropertyConfig, period: str, data_dir: str = "data") -> List[Path]:
    base = Path(data_dir) / cfg.property_code / "source_files"
    for candidate_period in list_periods(cfg, data_dir):
        if candidate_period > period:
            continue
        candidate_dir = base / candidate_period / "loan_statements"
        if candidate_dir.exists():
            pdfs = sorted(candidate_dir.glob("*.pdf"))
            if pdfs:
                return pdfs
    return []
