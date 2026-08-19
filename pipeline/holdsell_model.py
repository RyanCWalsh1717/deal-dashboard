"""Hold/Sell pro forma: assumption storage + Excel export.

The calculation logic (rollover schedule, debt amortization, IRR) lives
entirely in the exported workbook's own formulas — deliberately not
reimplemented here, so there is exactly one place the math can be wrong,
not two that could quietly disagree. This module only (1) persists the
assumption *values* Ryan edits in the app, and (2) writes those values,
plus whatever real data the app already has loaded, into a freshly-built
copy of that same workbook.

Sheet layout and formulas below build on the version audited on 2026-08-18
(this sandbox has no LibreOffice, so recalc.py can't run here — verification
is a manual formula-dump + defined-names cross-reference instead). Extended
2026-08-19 with a monthly cash flow tab, real OpEx-by-category, and debt
priced off a real SOFR curve + real per-tranche spread, per Ryan's review
of the first draft (referencing a real acquisition model his team already
uses, `210 Broadway v8.4.26.xlsm`, for the monthly/SOFR-curve conventions).
"""

from __future__ import annotations

import math
import re
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl
import yaml
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.workbook.workbook import Workbook

from pipeline.models import DistributionWorkbookResult, EntityTrialBalance, LoanStatement, RentRollResult
from pipeline.property_config import PropertyConfig

N_YEARS = 10  # projection years modeled; Hold Period (input) gates which of these actually count
YEAR_COLS = list(range(3, 3 + N_YEARS))  # columns C..L
N_MONTHS = 120  # 10 years x 12 — same window as N_YEARS, at monthly grain
MONTH_COLS = list(range(3, 3 + N_MONTHS))  # columns C..DR

# (key, label, section, kind) — kind is one of "years" | "pct" | "months" | "dollar" | "dollar_psf".
# Drives both the Streamlit form (grouped by section) and the Excel cell writes below — a single
# source of truth so the two never drift out of sync with each other.
ASSUMPTION_FIELDS = [
    ("hold_period", "Hold Period (years)", "Hold Period", "years"),
    ("rent_growth", "Rent Growth — Market Rent, annual %", "Inflation / Growth", "pct"),
    ("opex_growth", "Opex Growth, annual %", "Inflation / Growth", "pct"),
    ("tilc_growth", "TI / LC Growth, annual %", "Inflation / Growth", "pct"),
    ("capital_growth", "Capital Growth, annual %", "Inflation / Growth", "pct"),
    ("vacancy_credit", "General Vacancy & Credit Loss, % of collected rent", "Vacancy Credit", "pct"),
    ("renewal_downtime", "Downtime (months) — Renewal", "Market Leasing Assumptions", "months"),
    ("new_downtime", "Downtime (months) — New Lease", "Market Leasing Assumptions", "months"),
    ("renewal_free_rent", "Free Rent (months) — Renewal", "Market Leasing Assumptions", "months"),
    ("new_free_rent", "Free Rent (months) — New Lease", "Market Leasing Assumptions", "months"),
    ("renewal_ti", "TI ($/SF) — Renewal", "Market Leasing Assumptions", "dollar_psf"),
    ("new_ti", "TI ($/SF) — New Lease", "Market Leasing Assumptions", "dollar_psf"),
    ("renewal_lc", "LC ($/SF) — Renewal", "Market Leasing Assumptions", "dollar_psf"),
    ("new_lc", "LC ($/SF) — New Lease", "Market Leasing Assumptions", "dollar_psf"),
    ("renewal_prob", "Renewal Probability, %", "Market Leasing Assumptions", "pct"),
    ("market_rent_y1", "Market Rent, Year 1 ($/SF/Yr)", "Market Leasing Assumptions", "dollar_psf"),
    ("recovery_pct", "Recovery %, of Opex", "Market Leasing Assumptions", "pct"),
    ("cap_rate", "Exit Cap Rate, %", "Exit Assumptions", "pct"),
    ("cost_of_sale", "Cost of Sale, % of gross sale price", "Exit Assumptions", "pct"),
    ("refi_year", "Refinance Year (1-10; 0 = none modeled)", "Refinance Assumptions", "years"),
    ("refi_amount", "New Loan Amount ($)", "Refinance Assumptions", "dollar"),
    ("refi_rate", "New Loan Rate, %", "Refinance Assumptions", "pct"),
    ("refi_amort_years", "New Amortization (years)", "Refinance Assumptions", "years"),
    ("refi_io_years", "Interest-Only Period (years)", "Refinance Assumptions", "years"),
]

DEFAULT_ASSUMPTIONS: Dict[str, Optional[float]] = {key: None for key, *_ in ASSUMPTION_FIELDS}

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1A5C22")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14, color="1A5C22")
SECTION_FONT = Font(name=FONT_NAME, bold=True, size=11, color="1A5C22")
NOTE_FONT = Font(name=FONT_NAME, italic=True, size=9, color="666666")
INPUT_FONT = Font(name=FONT_NAME, color="0000FF")
FORMULA_FONT = Font(name=FONT_NAME, color="000000")
BOLD_FONT = Font(name=FONT_NAME, bold=True)
INPUT_FILL = PatternFill("solid", fgColor="FFFF00")
SECTION_FILL = PatternFill("solid", fgColor="E8F0E9")
THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MONEY_FMT = '$#,##0;($#,##0);"-"'
MONEY_PSF_FMT = '$#,##0.00;($#,##0.00);"-"'
PCT_FMT = "0.0%"
PCT1_FMT = "0.00%"
SF_FMT = "#,##0"
MULT_FMT = "0.00x"

# Real category-level OpEx totals from the Budget Comparison report — see
# pipeline/parsers/budget_comparison_report.py's OPEX_CATEGORY_LABELS for the
# exact source rows. Order here is display order in the model.
OPEX_CATEGORY_DISPLAY = [
    ("cleaning_janitorial", "Cleaning / Janitorial"),
    ("utilities", "Utilities"),
    ("general_repairs_maintenance", "General Repairs & Maintenance"),
    ("hvac_maintenance", "HVAC Maintenance"),
    ("plumbing", "Plumbing"),
    ("electrical_maintenance", "Electrical Maintenance"),
    ("security_fire_life_safety", "Security / Fire / Life Safety"),
    ("elevator_maintenance", "Elevator Maintenance"),
    ("landscaping", "Landscaping"),
    ("parking_garage_maintenance", "Parking & Garage Maintenance"),
    ("administrative", "Administrative"),
    ("insurance", "Insurance"),
    ("real_estate_taxes", "Real Estate Taxes"),
]


def _normalize_tranche(name: str) -> str:
    """Same convention as views/property_detail_view.py's _normalize_tranche
    (duplicated rather than imported — that one is private to the view
    module): strips spaces + trailing digits so "Note A" matches "Note A1"."""
    return re.sub(r"\d+$", "", name.lower().replace(" ", ""))


def _assumptions_path(cfg: PropertyConfig, data_dir: str = "data") -> Path:
    return Path(data_dir) / cfg.property_code / "holdsell_assumptions.yaml"


def load_assumptions(cfg: PropertyConfig, data_dir: str = "data") -> Dict[str, object]:
    path = _assumptions_path(cfg, data_dir)
    values = dict(DEFAULT_ASSUMPTIONS)
    values["sofr_curve"] = [None] * N_MONTHS
    if path.exists():
        with open(path, "r") as f:
            saved = yaml.safe_load(f) or {}
        for key in DEFAULT_ASSUMPTIONS:
            if key in saved:
                values[key] = saved[key]
        curve = saved.get("sofr_curve")
        if isinstance(curve, list) and len(curve) == N_MONTHS:
            values["sofr_curve"] = curve
    return values


def save_assumptions(cfg: PropertyConfig, assumptions: Dict[str, object], data_dir: str = "data") -> None:
    path = _assumptions_path(cfg, data_dir)
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w") as f:
        yaml.safe_dump(assumptions, f, default_flow_style=False, sort_keys=False)


def _style_header_row(ws, row, ncols, start_col=1):
    for c in range(start_col, start_col + ncols):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def _section_header(ws, row, text, span=2):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = SECTION_FONT
    for c in range(1, span + 1):
        ws.cell(row=row, column=c).fill = SECTION_FILL
    return row + 1


def _autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _input_cell(ws, row, col, value=None, fmt=None, comment=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = INPUT_FONT
    cell.fill = INPUT_FILL
    cell.border = BORDER
    if fmt:
        cell.number_format = fmt
    if comment:
        cell.comment = Comment(comment, "Deal Dashboard")
    return cell


def _real_cell(ws, row, col, value, fmt=None, source_comment=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = FORMULA_FONT
    cell.border = BORDER
    if fmt:
        cell.number_format = fmt
    if source_comment:
        cell.comment = Comment(source_comment, "Deal Dashboard")
    return cell


def _formula_cell(ws, row, col, formula, fmt=None, bold=False):
    cell = ws.cell(row=row, column=col, value=formula)
    cell.font = BOLD_FONT if bold else FORMULA_FONT
    cell.border = BORDER
    if fmt:
        cell.number_format = fmt
    return cell


def _year_header_row(ws, row, n_years=N_YEARS):
    ws.cell(row=row, column=1, value="Year")
    ws.cell(row=row, column=2, value="Year 0\n(Current)")
    for i, col in enumerate(YEAR_COLS, start=1):
        ws.cell(row=row, column=col, value=f"Year {i}")
    _style_header_row(ws, row, 2 + n_years, start_col=1)
    ws.row_dimensions[row].height = 26


def _month_header_row(ws, row, n_months=N_MONTHS):
    ws.cell(row=row, column=1, value="Month")
    for i, col in enumerate(MONTH_COLS, start=1):
        ws.cell(row=row, column=col, value=i)
    _style_header_row(ws, row, n_months, start_col=1)


def _col(c):
    return get_column_letter(c)


def _year_of_month(m: int) -> int:
    return (m - 1) // 12 + 1


def build_workbook(
    cfg: PropertyConfig,
    result: Optional[DistributionWorkbookResult],
    rent_roll: Optional[RentRollResult],
    loan_statements: List[LoanStatement],
    entity_trial_balances: List[EntityTrialBalance],
    assumptions: Dict[str, object],
    opex_categories: Optional[Dict[str, float]] = None,
) -> Workbook:
    if not result or not result.annual_budget_summary or not rent_roll:
        raise ValueError("build_workbook() needs a parsed distribution workbook (with annual budget summary) and rent roll.")

    a = dict(DEFAULT_ASSUMPTIONS)
    a["sofr_curve"] = [None] * N_MONTHS
    a.update({k: v for k, v in assumptions.items() if k in DEFAULT_ASSUMPTIONS or k == "sofr_curve"})
    opex_categories = opex_categories or {}

    annual = {l.account_code: l for l in result.annual_budget_summary.lines}
    as_of = rent_roll.as_of
    ytd_through = result.annual_budget_summary.period
    months_ytd = ytd_through.month if ytd_through else 12
    tb = entity_trial_balances[0] if entity_trial_balances else None

    def rollover_month(line):
        """Exact month (1-120) the suite's real lease expires, or Month 1 for
        vacant suites — average 30.4375 days/month, same conversion factor
        as the annual model's days/365.25 (x12), so the two always agree on
        which YEAR a suite rolls in even though this is month-precise."""
        if line.is_vacant or not line.lease_to or not as_of:
            return 1
        days = (line.lease_to - as_of).days
        months = days / 30.4375
        return max(1, min(N_MONTHS, math.ceil(months)))

    out = openpyxl.Workbook()
    out.remove(out.active)
    names: Dict[str, str] = {}

    def define(name, sheet, col, row):
        names[name] = f"'{sheet}'!${_col(col)}${row}"

    # =======================================================================
    # Sheet 1: Assumptions
    # =======================================================================
    wsA = out.create_sheet("Assumptions")
    wsA.sheet_view.showGridLines = False
    _autosize(wsA, [42, 20, 20, 55])

    wsA["A1"] = f"{cfg.display()} — Hold/Sell Model Assumptions"
    wsA["A1"].font = TITLE_FONT
    wsA["A2"] = (
        "Legend: yellow cells are your inputs, saved from the Deal Dashboard's Hold/Sell "
        "Assumptions page. Everything else is real data from the app's parsed source files, "
        "cited in each cell's comment."
    )
    wsA["A2"].font = NOTE_FONT
    wsA.merge_cells("A2:D2")
    wsA.row_dimensions[2].height = 28

    r = 4
    r = _section_header(wsA, r, "Hold Period", span=4)
    wsA.cell(row=r, column=1, value="Hold Period (years)")
    _input_cell(wsA, r, 2, value=a["hold_period"], fmt="0", comment="How many of the 10 projection years count toward the exit/IRR calc.")
    define("HoldPeriod", "Assumptions", 2, r)
    r += 2

    r = _section_header(wsA, r, "Inflation / Growth Assumptions", span=4)
    for label, name, key in [
        ("Rent Growth — Market Rent, annual %", "RentGrowth", "rent_growth"),
        ("Opex Growth, annual %", "OpexGrowth", "opex_growth"),
        ("TI / LC Growth, annual %", "TILCGrowth", "tilc_growth"),
        ("Capital Growth, annual %", "CapitalGrowth", "capital_growth"),
    ]:
        wsA.cell(row=r, column=1, value=label)
        _input_cell(wsA, r, 2, value=a[key], fmt=PCT_FMT)
        define(name, "Assumptions", 2, r)
        r += 1
    r += 1

    r = _section_header(wsA, r, "Vacancy Credit", span=4)
    wsA.cell(row=r, column=1, value="General Vacancy & Credit Loss, % of collected rent")
    _input_cell(wsA, r, 2, value=a["vacancy_credit"], fmt=PCT_FMT, comment=(
        "Frictional vacancy/bad-debt allowance applied on TOP of the explicit suite-level "
        "downtime already modeled on the Rollover sheet — this is the general cushion, not the "
        "known rollover gaps."
    ))
    define("VacancyCredit", "Assumptions", 2, r)
    r += 2

    r = _section_header(wsA, r, "Market Leasing Assumptions", span=4)
    wsA.cell(row=r, column=1, value="Applied at every suite's rollover — blended by Renewal Probability below.").font = NOTE_FONT
    r += 1
    hdr_row = r
    for i, h in enumerate(["", "Renewal", "New Lease (post-downtime)"], start=1):
        wsA.cell(row=hdr_row, column=i, value=h)
    _style_header_row(wsA, hdr_row, 3)
    r += 1
    for label, key_suffix, ren_key, new_key in [
        ("Downtime (months)", "Downtime", "renewal_downtime", "new_downtime"),
        ("Free Rent (months)", "FreeRent", "renewal_free_rent", "new_free_rent"),
        ("TI ($/SF)", "TI", "renewal_ti", "new_ti"),
        ("LC ($/SF)", "LC", "renewal_lc", "new_lc"),
    ]:
        wsA.cell(row=r, column=1, value=label)
        fmt = MONEY_PSF_FMT if key_suffix in ("TI", "LC") else "0.0"
        _input_cell(wsA, r, 2, value=a[ren_key], fmt=fmt)
        _input_cell(wsA, r, 3, value=a[new_key], fmt=fmt)
        define(f"Renewal{key_suffix}", "Assumptions", 2, r)
        define(f"New{key_suffix}", "Assumptions", 3, r)
        r += 1
    wsA.cell(row=r, column=1, value="Renewal Probability, %")
    _input_cell(wsA, r, 2, value=a["renewal_prob"], fmt=PCT_FMT, comment="Probability an expiring/vacant suite renews vs. re-leases as a new lease (post-downtime).")
    define("RenewalProb", "Assumptions", 2, r)
    r += 1
    wsA.cell(row=r, column=1, value="Market Rent, Year 1 ($/SF/Yr)")
    _input_cell(wsA, r, 2, value=a["market_rent_y1"], fmt=MONEY_PSF_FMT, comment=(
        "Market rent for a suite rolling over in Year 1. Grows by Rent Growth each year after "
        "that — applies to both Renewal and New Lease outcomes; both reset to market, they "
        "differ in downtime/free rent/TI/LC."
    ))
    define("MarketRentY1", "Assumptions", 2, r)
    r += 1
    wsA.cell(row=r, column=1, value="Recovery %, of Opex")
    _input_cell(wsA, r, 2, value=a["recovery_pct"], fmt=PCT_FMT, comment=(
        "Blended recovery structure — % of operating expenses recouped from tenants as "
        "reimbursement revenue. Simplification: applied uniformly rather than lease-by-lease; "
        "scaled down in rollover years by the same rent-collection factor as the suite's own "
        "downtime (see Cash Flow Projection)."
    ))
    define("RecoveryPct", "Assumptions", 2, r)
    r += 2

    r = _section_header(wsA, r, "Exit Assumptions", span=4)
    wsA.cell(row=r, column=1, value="Exit Cap Rate, %")
    _input_cell(wsA, r, 2, value=a["cap_rate"], fmt=PCT_FMT, comment="Applied to the exit year's own (trailing) NOI — not forward NOI.")
    define("CapRate", "Assumptions", 2, r)
    r += 1
    wsA.cell(row=r, column=1, value="Cost of Sale, % of gross sale price")
    _input_cell(wsA, r, 2, value=a["cost_of_sale"], fmt=PCT_FMT)
    define("CostOfSale", "Assumptions", 2, r)
    r += 2

    r = _section_header(wsA, r, "Refinance Assumptions", span=4)
    wsA.cell(row=r, column=1, value="Refinance Year (1-10; 0 or blank = no refinance modeled)")
    _input_cell(wsA, r, 2, value=a["refi_year"], fmt="0")
    define("RefiYear", "Assumptions", 2, r)
    r += 1
    wsA.cell(row=r, column=1, value="New Loan Amount ($)")
    _input_cell(wsA, r, 2, value=a["refi_amount"], fmt=MONEY_FMT)
    define("RefiAmount", "Assumptions", 2, r)
    r += 1
    wsA.cell(row=r, column=1, value="New Loan Rate, %")
    _input_cell(wsA, r, 2, value=a["refi_rate"], fmt=PCT_FMT)
    define("RefiRate", "Assumptions", 2, r)
    r += 1
    wsA.cell(row=r, column=1, value="New Amortization (years)")
    _input_cell(wsA, r, 2, value=a["refi_amort_years"], fmt="0")
    define("RefiAmortYears", "Assumptions", 2, r)
    r += 1
    wsA.cell(row=r, column=1, value="Interest-Only Period (years)")
    _input_cell(wsA, r, 2, value=a["refi_io_years"], fmt="0")
    define("RefiIOYears", "Assumptions", 2, r)
    r += 1
    wsA.cell(row=r, column=1, value="Refinance Month (computed)")
    _formula_cell(wsA, r, 2, "=IF(RefiYear>0,(RefiYear-1)*12+1,0)", "0")
    wsA.cell(row=r, column=2).comment = Comment(
        "First month of the Refinance Year — the whole of that year is modeled as already on "
        "the new loan, same convention the annual model used.", "Deal Dashboard")
    define("RefiMonth", "Assumptions", 2, r)
    r += 2

    # ---- Real debt data: per-tranche balance + real spread + real SOFR ----
    r = _section_header(wsA, r, "Current Debt — Real Data", span=5)
    wsA.cell(row=r, column=1, value=(
        "Spread is real (from the distribution workbook's debt tab). Current SOFR below is "
        "derived from your real loan statements (all-in rate minus spread) rather than the "
        "distribution workbook's own SOFR figure, which is a forecast/stale snapshot — same "
        "forecast-vs-actual distinction as the rest of this app's debt figures."
    )).font = NOTE_FONT
    wsA.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    wsA.row_dimensions[r].height = 40
    r += 1
    hdr_row = r
    for i, h in enumerate(["Tranche", "Balance", "Spread", "Real All-In (loan stmt)", "Implied Current SOFR"], start=1):
        wsA.cell(row=hdr_row, column=i, value=h)
    _style_header_row(wsA, hdr_row, 5)
    r += 1

    debt_by_norm = {
        _normalize_tranche(t.tranche_name): t for t in (result.debt.tranches if result.debt else [])
    }
    tranche_rows = []  # (row, real_balance, real_spread)
    implied_sofr_rows = []
    for stmt in loan_statements:
        matched = debt_by_norm.get(_normalize_tranche(stmt.tranche_name))
        spread = matched.interest_rate if matched else None
        wsA.cell(row=r, column=1, value=stmt.tranche_name).border = BORDER
        _real_cell(wsA, r, 2, stmt.principal_balance, MONEY_FMT, "Source: loan servicer statement, currently loaded period.")
        _real_cell(wsA, r, 3, spread, PCT1_FMT, "Source: distribution workbook debt tab (real spread over SOFR).")
        _real_cell(wsA, r, 4, stmt.interest_rate, PCT1_FMT, "Source: loan servicer statement (real all-in rate).")
        if spread is not None:
            _formula_cell(wsA, r, 5, f"=D{r}-C{r}", PCT1_FMT)
            implied_sofr_rows.append(r)
        define(f"Tranche{len(tranche_rows) + 1}Balance", "Assumptions", 2, r)
        define(f"Tranche{len(tranche_rows) + 1}Spread", "Assumptions", 3, r)
        tranche_rows.append((r, stmt.principal_balance, spread))
        r += 1
    n_tranches = len(tranche_rows)
    if not loan_statements:
        wsA.cell(row=r, column=1, value="(no loan statements loaded)").font = NOTE_FONT
        r += 1
    r += 1

    wsA.cell(row=r, column=1, value="Current SOFR (shared, real, derived from loan statements)")
    if implied_sofr_rows:
        _formula_cell(wsA, r, 2, f"=AVERAGE({','.join(f'E{rr}' for rr in implied_sofr_rows)})", PCT1_FMT, bold=True)
    else:
        _formula_cell(wsA, r, 2, "=0", PCT1_FMT, bold=True)
    define("CurrentSOFR", "Assumptions", 2, r)
    r += 1
    wsA.cell(row=r, column=1, value="SOFR per Distribution Workbook (forecast reference only, not used in calcs)")
    _real_cell(
        wsA, r, 2, result.debt.sofr_as_of if result.debt else None, PCT1_FMT,
        "Source: distribution workbook debt tab. Shown for comparison only — it's a "
        "forecast/stale snapshot, not necessarily today's real SOFR; the model uses the "
        "derived 'Current SOFR' above instead."
    )
    r += 2

    r += 2

    contributions = tb.contributions if tb else None
    distributions = tb.distributions if tb else None
    wsA.cell(row=r, column=1, value=f"Contributions to Date ({tb.entity_code if tb else 'n/a'}, real)")
    _real_cell(wsA, r, 2, contributions or 0, MONEY_FMT, "Source: trial balance, currently loaded period.")
    contrib_row = r
    r += 1
    wsA.cell(row=r, column=1, value=f"Distributions to Date ({tb.entity_code if tb else 'n/a'}, real)")
    _real_cell(wsA, r, 2, distributions or 0, MONEY_FMT, "Source: trial balance, currently loaded period.")
    distrib_row = r
    r += 1
    wsA.cell(row=r, column=1, value="Starting Net Invested Equity (Contributions − Distributions)")
    _formula_cell(wsA, r, 2, f"=B{contrib_row}-B{distrib_row}", MONEY_FMT, bold=True)
    wsA.cell(row=r, column=2).comment = Comment(
        "This is the loaded entity's own net invested capital, not necessarily what a specific "
        "LP/GP holds after the JV waterfall — a reasonable real starting point for a top-line IRR, "
        "but override if you want the actual equity basis for a specific investor.", "Deal Dashboard"
    )
    define("StartingEquity", "Assumptions", 2, r)
    r += 2

    r = _section_header(wsA, r, "Blended Market Leasing Terms (computed)", span=4)
    for label, key_suffix in [
        ("Blended Downtime (months)", "Downtime"),
        ("Blended Free Rent (months)", "FreeRent"),
        ("Blended TI ($/SF)", "TI"),
        ("Blended LC ($/SF)", "LC"),
    ]:
        wsA.cell(row=r, column=1, value=label)
        fmt = MONEY_PSF_FMT if key_suffix in ("TI", "LC") else "0.0"
        _formula_cell(wsA, r, 2, f"=RenewalProb*Renewal{key_suffix}+(1-RenewalProb)*New{key_suffix}", fmt)
        define(f"Blended{key_suffix}", "Assumptions", 2, r)
        r += 1
    r += 1

    # ---- Real OpEx category baselines ----
    r = _section_header(wsA, r, "Operating Expense Categories — Real Annual BUDGET Baseline", span=3)
    wsA.cell(row=r, column=1, value=(
        "Each category below is this year's full-year BUDGET (not actual) — the Budget Comparison "
        "report doesn't carry actual figures for months that haven't happened yet, so a full-year "
        "figure has to come from budget. Year 1 of the Monthly/Cash Flow sheets grows off these "
        "budget baselines; Year 0 (elsewhere in this model) still uses the YTD ACTUAL run-rate, "
        "same as every other Year 0 figure — so a real vs. budget variance in the current year will "
        "show up as a step between Year 0 and Year 1, not a smooth trend."
    )).font = NOTE_FONT
    wsA.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    wsA.row_dimensions[r].height = 55
    r += 1
    hdr_row = r
    for i, h in enumerate(["Category", "Annual Budget Baseline (real)"], start=1):
        wsA.cell(row=hdr_row, column=i, value=h)
    _style_header_row(wsA, hdr_row, 2)
    r += 1
    opex_category_rows: Dict[str, int] = {}
    for key, label in OPEX_CATEGORY_DISPLAY:
        wsA.cell(row=r, column=1, value=label).border = BORDER
        value = opex_categories.get(key)
        _real_cell(wsA, r, 2, value, MONEY_FMT, "Source: Yardi Budget Comparison report, currently loaded period (Annual column — this year's full-year BUDGET, not actual).")
        opex_category_rows[key] = r
        r += 1
    r += 1

    # ---- Monthly SOFR curve (input row, 120 months) ----
    r = _section_header(wsA, r, "SOFR Curve (Monthly)", span=4)
    wsA.cell(row=r, column=1, value=(
        "Optional: type/paste your own real forward SOFR curve into any month below (e.g. from "
        "a market-data source, the same way your team's other models do it). Any month left "
        "blank falls back to Current SOFR above."
    )).font = NOTE_FONT
    wsA.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    wsA.row_dimensions[r].height = 40
    r += 1
    _month_header_row(wsA, r)
    r += 1
    sofr_curve_row = r
    saved_curve = a.get("sofr_curve") or [None] * N_MONTHS
    for i, col in enumerate(MONTH_COLS):
        curve_value = saved_curve[i] if i < len(saved_curve) else None
        _input_cell(wsA, r, col, value=curve_value, fmt=PCT1_FMT)
    wsA.cell(row=r, column=1, value="SOFR Curve")
    define("SofrCurveRow", "Assumptions", 3, r)  # row number only needed; column varies per month

    wsA.freeze_panes = "C4"

    def sofr_used_formula(month_num: int) -> str:
        cl = _col(MONTH_COLS[month_num - 1])
        return f'IF(Assumptions!{cl}{sofr_curve_row}="",CurrentSOFR,Assumptions!{cl}{sofr_curve_row})'

    # =======================================================================
    # Sheet 2: Rent Roll (Current)
    # =======================================================================
    ws1 = out.create_sheet("Rent Roll (Current)")
    ws1.sheet_view.showGridLines = False
    _autosize(ws1, [10, 12, 10, 12, 12, 30, 16, 12, 14, 12])

    ws1["A1"] = f"{cfg.display()} — Current Rent Roll"
    ws1["A1"].font = TITLE_FONT
    ws1["A2"] = f"As of {as_of.strftime('%B %d, %Y') if as_of else 'n/a'} — source: Deal Dashboard rent roll, currently loaded period."
    ws1["A2"].font = NOTE_FONT

    r = 4
    header_row = r
    headers = ["Building", "Floor", "Unit", "SF", "Status", "Tenant", "Annual Rent", "Rent PSF", "Lease Expiration", "Rollover Year"]
    for i, h in enumerate(headers, start=1):
        ws1.cell(row=header_row, column=i, value=h)
    _style_header_row(ws1, header_row, len(headers))
    r += 1
    rent_roll_rows = {}
    for idx, line in enumerate(rent_roll.lines):
        rm = rollover_month(line)
        ry = _year_of_month(rm)
        ws1.cell(row=r, column=1, value=line.building).border = BORDER
        ws1.cell(row=r, column=2, value=str(line.floor)).border = BORDER
        ws1.cell(row=r, column=3, value=line.unit_code).border = BORDER
        c4 = ws1.cell(row=r, column=4, value=line.unit_area); c4.number_format = SF_FMT; c4.border = BORDER
        ws1.cell(row=r, column=5, value="VACANT" if line.is_vacant else "Leased").border = BORDER
        ws1.cell(row=r, column=6, value=line.tenant_name or "").border = BORDER
        c7 = ws1.cell(row=r, column=7, value=line.annual_rent); c7.number_format = MONEY_FMT; c7.border = BORDER
        c8 = ws1.cell(row=r, column=8, value=line.annual_rent_psf); c8.number_format = MONEY_PSF_FMT; c8.border = BORDER
        ws1.cell(row=r, column=9, value=line.lease_to.strftime("%m/%d/%Y") if line.lease_to else "").border = BORDER
        ryc = ws1.cell(row=r, column=10, value=ry); ryc.border = BORDER
        ryc.comment = Comment(
            f"Computed from the real lease expiration date — rolls over in Month {rm} (Year {ry}). "
            "Vacant suites = Month 1. Simplification: models each suite's FIRST rollover only.",
            "Deal Dashboard",
        )
        for c in range(1, 11):
            ws1.cell(row=r, column=c).font = FORMULA_FONT
        rent_roll_rows[idx] = (r, line, ry, rm)
        r += 1
    ws1.freeze_panes = "A5"

    # =======================================================================
    # Sheet 3: Rollover & Rent Projection (annual, unchanged mechanics)
    # =======================================================================
    ws3 = out.create_sheet("Rollover & Rent Projection")
    ws3.sheet_view.showGridLines = False
    _autosize(ws3, [34] + [14] * (N_YEARS + 1))

    ws3["A1"] = "Rollover & Rent Projection"
    ws3["A1"].font = TITLE_FONT
    ws3["A2"] = (
        "Pre-rollover years hold each suite's real current rent flat. At rollover, rent resets to "
        "the escalated market rate below; downtime/free rent reduce cash collected only in the "
        "rollover year itself. The Monthly Cash Flow sheet levels each year's total evenly across "
        "its 12 months, except TI/LC which fires in the exact rollover month."
    )
    ws3["A2"].font = NOTE_FONT
    ws3.merge_cells("A2:F2")
    ws3.row_dimensions[2].height = 40

    r = 4
    _year_header_row(ws3, r)
    r += 1

    ws3.cell(row=r, column=1, value="Market Rent, $/SF/Yr (escalated)").font = BOLD_FONT
    market_rent_row = r
    _formula_cell(ws3, r, 3, "=MarketRentY1", MONEY_PSF_FMT)
    for i, col in enumerate(YEAR_COLS[1:], start=1):
        prev_col = YEAR_COLS[i - 1]
        _formula_cell(ws3, r, col, f"={_col(prev_col)}{r}*(1+RentGrowth)", MONEY_PSF_FMT)
    r += 2

    ws3.cell(row=r, column=1, value="Per-Suite Contract Rent ($/Yr)").font = SECTION_FONT
    r += 1
    contract_rows = {}
    cash_rows = {}
    for idx, (rr_row, line, ry, rm) in rent_roll_rows.items():
        label = f"{line.building}-{line.floor}-{line.unit_code}" + (" (vacant)" if line.is_vacant else f" — {line.tenant_name}")
        ws3.cell(row=r, column=1, value=f"Contract Rent: {label}")
        sf_ref = f"'Rent Roll (Current)'!D{rr_row}"
        cur_rent_ref = f"'Rent Roll (Current)'!G{rr_row}"
        for col in YEAR_COLS:
            year_num = col - YEAR_COLS[0] + 1
            cl = _col(col)
            formula = f"={cur_rent_ref}" if year_num < ry else f"={sf_ref}*{cl}${market_rent_row}"
            _formula_cell(ws3, r, col, formula, MONEY_FMT)
        contract_rows[idx] = r
        r += 1
    r += 1

    ws3.cell(row=r, column=1, value="Per-Suite Cash Rent ($/Yr, net of rollover downtime & free rent)").font = SECTION_FONT
    r += 1
    for idx, (rr_row, line, ry, rm) in rent_roll_rows.items():
        label = f"{line.building}-{line.floor}-{line.unit_code}" + (" (vacant)" if line.is_vacant else f" — {line.tenant_name}")
        ws3.cell(row=r, column=1, value=f"Cash Rent: {label}")
        crow = contract_rows[idx]
        for col in YEAR_COLS:
            year_num = col - YEAR_COLS[0] + 1
            cl = _col(col)
            if year_num == ry:
                formula = f"={cl}{crow}*(12-BlendedDowntime-BlendedFreeRent)/12"
            else:
                formula = f"={cl}{crow}"
            _formula_cell(ws3, r, col, formula, MONEY_FMT)
        cash_rows[idx] = r
        r += 1
    r += 1

    ws3.cell(row=r, column=1, value="Total Contract Rent").font = BOLD_FONT
    total_contract_row = r
    for col in YEAR_COLS:
        cl = _col(col)
        _formula_cell(ws3, r, col, "=" + "+".join(f"{cl}{cr}" for cr in contract_rows.values()), MONEY_FMT, bold=True)
    r += 1
    ws3.cell(row=r, column=1, value="Total Cash Rent (Gross Potential Rent)").font = BOLD_FONT
    total_cash_row = r
    for col in YEAR_COLS:
        cl = _col(col)
        _formula_cell(ws3, r, col, "=" + "+".join(f"{cl}{cr}" for cr in cash_rows.values()), MONEY_FMT, bold=True)
    r += 1
    ws3.cell(row=r, column=1, value="Rent Collection Factor (Cash / Contract)")
    ws3.cell(row=r, column=1).comment = Comment(
        "Used to scale Recovery Income proportionally in rollover years.", "Deal Dashboard")
    collection_factor_row = r
    for col in YEAR_COLS:
        cl = _col(col)
        _formula_cell(ws3, r, col, f"={cl}{total_cash_row}/{cl}{total_contract_row}", PCT_FMT)
    r += 2

    ws3.cell(row=r, column=1, value="TI / LC Cost This Year ($)").font = SECTION_FONT
    ws3.cell(row=r, column=1).comment = Comment(
        "Fires only in each suite's own rollover year: SF x (Blended TI + Blended LC), both "
        "$/SF. Escalates by TI/LC Growth for rollovers happening in later years. The Monthly "
        "sheet places this same dollar amount in the suite's exact rollover month.", "Deal Dashboard")
    r += 1
    ti_lc_row_start = r
    ti_lc_rows_by_suite: Dict[int, int] = {}
    for idx, (rr_row, line, ry, rm) in rent_roll_rows.items():
        label = f"{line.building}-{line.floor}-{line.unit_code}"
        ws3.cell(row=r, column=1, value=f"TI/LC: {label}")
        sf_ref = f"'Rent Roll (Current)'!D{rr_row}"
        for col in YEAR_COLS:
            year_num = col - YEAR_COLS[0] + 1
            if year_num == ry:
                escal = f"*(1+TILCGrowth)^{ry - 1}"
                formula = f"=({sf_ref}*(BlendedTI+BlendedLC){escal})"
            else:
                formula = "=0"
            _formula_cell(ws3, r, col, formula, MONEY_FMT)
        ti_lc_rows_by_suite[idx] = r
        r += 1
    ti_lc_row_end = r - 1
    ws3.cell(row=r, column=1, value="Total TI / LC Cost").font = BOLD_FONT
    total_ti_lc_row = r
    for col in YEAR_COLS:
        cl = _col(col)
        _formula_cell(ws3, r, col, f"=SUM({cl}{ti_lc_row_start}:{cl}{ti_lc_row_end})", MONEY_FMT, bold=True)

    ws3.freeze_panes = "B5"

    # =======================================================================
    # Sheet 4: Monthly Cash Flow (new)
    # =======================================================================
    ws_m = out.create_sheet("Monthly Cash Flow")
    ws_m.sheet_view.showGridLines = False
    _autosize(ws_m, [40] + [11] * N_MONTHS)

    ws_m["A1"] = "Monthly Cash Flow"
    ws_m["A1"].font = TITLE_FONT
    ws_m["A2"] = (
        "Full 10-year (120-month) detail regardless of Hold Period. Rent/OpEx/CapEx are each "
        "year's annual figure leveled evenly across its 12 months (a documented simplification — "
        "only intra-year timing is leveled, not which year things happen); TI/LC and the debt "
        "schedule are true month-precise."
    )
    ws_m["A2"].font = NOTE_FONT
    ws_m.merge_cells("A2:F2")
    ws_m.row_dimensions[2].height = 40

    r = 4
    _month_header_row(ws_m, r)
    r += 2

    def month_row(label, formula_fn, fmt=MONEY_FMT, bold=False):
        nonlocal r
        ws_m.cell(row=r, column=1, value=label)
        if bold:
            ws_m.cell(row=r, column=1).font = BOLD_FONT
        for m, col in enumerate(MONTH_COLS, start=1):
            f = formula_fn(m, col)
            if f is not None:
                _formula_cell(ws_m, r, col, f, fmt, bold=bold)
        this_row = r
        r += 1
        return this_row

    def year_col_for_month(m: int) -> int:
        return YEAR_COLS[_year_of_month(m) - 1]

    gpr_m_row = month_row(
        "Gross Potential Rent (Total Cash Rent)",
        lambda m, col: f"='Rollover & Rent Projection'!{_col(year_col_for_month(m))}{total_cash_row}/12",
    )
    vac_credit_m_row = month_row(
        "Less: Vacancy & Credit Loss",
        lambda m, col: f"=-{_col(col)}{gpr_m_row}*VacancyCredit",
    )
    egi_m_row = month_row(
        "Effective Rental Income",
        lambda m, col: f"={_col(col)}{gpr_m_row}+{_col(col)}{vac_credit_m_row}", bold=True,
    )

    # OpEx by category — each category's real annual baseline, escalated once per calendar
    # year (flat within the year, same step convention as the annual model), divided by 12.
    opex_cat_m_rows: Dict[str, int] = {}
    for key, label in OPEX_CATEGORY_DISPLAY:
        baseline_row = opex_category_rows[key]

        def _cat_formula(m, col, baseline_row=baseline_row):
            year_exp = _year_of_month(m) - 1
            growth = f"*(1+OpexGrowth)^{year_exp}" if year_exp else ""
            return f"=-Assumptions!B{baseline_row}{growth}/12"

        opex_cat_m_rows[key] = month_row(label, _cat_formula)
    total_opex_m_row = month_row(
        "Total Operating Expenses",
        lambda m, col: "=" + "+".join(f"{_col(col)}{rr}" for rr in opex_cat_m_rows.values()), bold=True,
    )

    nonop_actual = annual["non_operating_expenses"].actual_value if annual.get("non_operating_expenses") else None
    nonop_m_row = month_row(
        "Non-operating Expenses",
        lambda m, col: (
            f"=-{nonop_actual or 0}/{months_ytd}*12*(1+OpexGrowth)^{_year_of_month(m) - 1}/12"
        ),
    )
    recovery_m_row = month_row(
        "Recovery Income",
        lambda m, col: (
            f"=RecoveryPct*-{_col(col)}{total_opex_m_row}*"
            f"'Rollover & Rent Projection'!{_col(year_col_for_month(m))}{collection_factor_row}"
        ),
    )
    total_rev_m_row = month_row(
        "Total Revenue",
        lambda m, col: f"={_col(col)}{egi_m_row}+{_col(col)}{recovery_m_row}", bold=True,
    )
    noi_m_row = month_row(
        "NOI",
        lambda m, col: f"={_col(col)}{total_rev_m_row}+{_col(col)}{total_opex_m_row}+{_col(col)}{nonop_m_row}",
        bold=True,
    )

    ti_lc_m_rows_by_suite: Dict[int, int] = {}
    for idx, (rr_row, line, ry, rm) in rent_roll_rows.items():
        annual_ti_lc_row = ti_lc_rows_by_suite[idx]
        label = f"{line.building}-{line.floor}-{line.unit_code}"

        def _ti_lc_formula(m, col, rm=rm, annual_ti_lc_row=annual_ti_lc_row):
            if m != rm:
                return "=0"
            return f"='Rollover & Rent Projection'!{_col(year_col_for_month(m))}{annual_ti_lc_row}"

        ti_lc_m_rows_by_suite[idx] = month_row(f"TI/LC: {label}", _ti_lc_formula)
    total_ti_lc_m_row = month_row(
        "Total TI / LC Cost", lambda m, col: "=" + "+".join(f"{_col(col)}{rr}" for rr in ti_lc_m_rows_by_suite.values()),
        bold=True,
    )
    less_ti_lc_m_row = month_row(
        "Less: TI / LC", lambda m, col: f"=-{_col(col)}{total_ti_lc_m_row}",
    )

    capex_budget = annual["capital_expenditures"].budget_value if annual.get("capital_expenditures") else None
    capex_m_row = month_row(
        "Less: Capital Expenditures",
        lambda m, col: f"=-{capex_budget or 0}*(1+CapitalGrowth)^{_year_of_month(m) - 1}/12",
    )
    cfbds_m_row = month_row(
        "Cash Flow Before Debt Service",
        lambda m, col: f"={_col(col)}{noi_m_row}+{_col(col)}{less_ti_lc_m_row}+{_col(col)}{capex_m_row}",
        bold=True,
    )

    r += 1
    ws_m.cell(row=r, column=1, value="Debt Schedule (real per-tranche balance + SOFR curve + real spread)").font = SECTION_FONT
    r += 1
    ws_m.cell(row=r, column=1, value=(
        "Existing tranches are modeled interest-only through refinance (flat balance; real debt "
        "service ties closely to pure interest, see Assumptions) — each priced off the SOFR "
        "curve plus its own real spread. The new post-refinance loan amortizes monthly at its "
        "own fixed rate via PMT/PPMT."
    )).font = NOTE_FONT
    ws_m.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws_m.row_dimensions[r].height = 40
    r += 1

    old_tranche_balance_rows = []
    old_tranche_interest_rows = []
    for i, (trow, real_balance, real_spread) in enumerate(tranche_rows, start=1):
        tranche_label = loan_statements[i - 1].tranche_name if i - 1 < len(loan_statements) else f"Tranche {i}"
        bal_row = month_row(
            f"Balance: {tranche_label}",
            lambda m, col, i=i: f"=IF(AND(RefiYear>0,{m}>=RefiMonth),0,Tranche{i}Balance)",
        )
        int_row = month_row(
            f"Interest: {tranche_label}",
            lambda m, col, i=i, bal_row=bal_row: (
                f"={_col(col)}{bal_row}*({sofr_used_formula(m)}+Tranche{i}Spread)/12"
            ),
        )
        old_tranche_balance_rows.append(bal_row)
        old_tranche_interest_rows.append(int_row)

    # Beginning(m) references Ending(m-1) — a forward reference to a row that doesn't exist
    # yet at this point, since Ending is written further down. Written as "=0" here and
    # overwritten in the backfill pass right after New Loan: Ending Balance is defined below.
    new_loan_beginning_row = month_row(
        "New Loan (Post-Refi): Beginning Balance",
        lambda m, col: "=0",
    )
    new_loan_months_since_row = month_row(
        "New Loan: Months Since Refi",
        lambda m, col: f"=IF({m}>=RefiMonth,{m}-RefiMonth+1,0)", fmt="0",
    )
    new_loan_io_flag_row = month_row(
        "New Loan: Interest-Only This Month?",
        lambda m, col: f"=IF({_col(col)}{new_loan_months_since_row}>0,IF({_col(col)}{new_loan_months_since_row}<=RefiIOYears*12,1,0),0)",
        fmt="0",
    )
    new_loan_interest_row = month_row(
        "New Loan: Interest",
        lambda m, col: f"={_col(col)}{new_loan_beginning_row}*RefiRate/12",
    )
    new_loan_principal_row = month_row(
        "New Loan: Principal",
        lambda m, col: (
            f"=IF({_col(col)}{new_loan_io_flag_row}=1,0,IF({m}>=RefiMonth,"
            f"PPMT(RefiRate/12,{_col(col)}{new_loan_months_since_row}-RefiIOYears*12,"
            f"RefiAmortYears*12,-RefiAmount),0))"
        ),
    )
    new_loan_ending_row = month_row(
        "New Loan: Ending Balance",
        lambda m, col: f"={_col(col)}{new_loan_beginning_row}-{_col(col)}{new_loan_principal_row}",
    )
    # Backfill New Loan Beginning Balance now that its own Ending row's position is known —
    # written as a fresh pass since Beginning(m) references Ending(m-1), a forward reference
    # at the time Beginning was first written.
    for m, col in enumerate(MONTH_COLS, start=1):
        prev_cl = _col(col - 1)
        formula = (
            f"=IF(RefiYear=0,0,IF({m}<RefiMonth,0,IF({m}=RefiMonth,RefiAmount,"
            f"{prev_cl}{new_loan_ending_row})))"
        )
        _formula_cell(ws_m, new_loan_beginning_row, col, formula, MONEY_FMT)

    total_interest_m_row = month_row(
        "Total Interest",
        lambda m, col: "=" + "+".join(f"{_col(col)}{rr}" for rr in old_tranche_interest_rows) + f"+{_col(col)}{new_loan_interest_row}",
        bold=True,
    )
    total_principal_m_row = month_row(
        "Total Principal", lambda m, col: f"={_col(col)}{new_loan_principal_row}", bold=True,
    )
    total_ds_m_row = month_row(
        "Total Debt Service",
        lambda m, col: f"={_col(col)}{total_interest_m_row}+{_col(col)}{total_principal_m_row}", bold=True,
    )
    total_beginning_balance_m_row = month_row(
        "Total Beginning Balance (all tranches)",
        lambda m, col: "=" + "+".join(f"{_col(col)}{rr}" for rr in old_tranche_balance_rows) + f"+{_col(col)}{new_loan_beginning_row}",
        bold=True,
    )
    total_ending_balance_m_row = month_row(
        "Total Ending Balance (all tranches)",
        lambda m, col: "=" + "+".join(f"{_col(col)}{rr}" for rr in old_tranche_balance_rows) + f"+{_col(col)}{new_loan_ending_row}",
        bold=True,
    )
    levered_cf_m_row = month_row(
        "Levered Cash Flow (before sale)",
        lambda m, col: f"={_col(col)}{cfbds_m_row}-{_col(col)}{total_ds_m_row}", bold=True,
    )

    ws_m.freeze_panes = "B5"

    # =======================================================================
    # Sheet 5: Cash Flow Projection (annual) — now a rollup of Monthly
    # =======================================================================
    ws4 = out.create_sheet("Cash Flow Projection")
    ws4.sheet_view.showGridLines = False
    _autosize(ws4, [40] + [14] * (N_YEARS + 1))

    ws4["A1"] = "Cash Flow Projection (Annual Summary)"
    ws4["A1"].font = TITLE_FONT
    ws4["A2"] = (
        "Sums each year's 12 months from the Monthly Cash Flow sheet (balance-type rows take "
        "the specific month instead of summing). Shows all 10 modeled years regardless of Hold "
        "Period — the Returns sheet applies the Hold Period gate and the sale."
    )
    ws4["A2"].font = NOTE_FONT
    ws4.merge_cells("A2:F2")
    ws4.row_dimensions[2].height = 28

    r = 4
    _year_header_row(ws4, r)
    r += 2

    def annual_flow_row(label, month_row_ref, fmt=MONEY_FMT, bold=False, year0_formula=None):
        """Annual value = SUM of that year's 12 months from the given Monthly-sheet row."""
        nonlocal r
        ws4.cell(row=r, column=1, value=label)
        if bold:
            ws4.cell(row=r, column=1).font = BOLD_FONT
        if year0_formula is not None:
            _formula_cell(ws4, r, 2, year0_formula, fmt, bold=bold)
        for yi, ycol in enumerate(YEAR_COLS, start=1):
            months_this_year = [MONTH_COLS[m - 1] for m in range((yi - 1) * 12 + 1, yi * 12 + 1)]
            first_cl, last_cl = _col(months_this_year[0]), _col(months_this_year[-1])
            formula = f"=SUM('Monthly Cash Flow'!{first_cl}{month_row_ref}:{last_cl}{month_row_ref})"
            _formula_cell(ws4, r, ycol, formula, fmt, bold=bold)
        this_row = r
        r += 1
        return this_row

    def annual_point_row(label, month_row_ref, take="last", fmt=MONEY_FMT, bold=False, year0_formula=None):
        """Annual value = the specific month's value (first or last month of the year) from
        the Monthly sheet — for balance-type rows that shouldn't be summed across the year."""
        nonlocal r
        ws4.cell(row=r, column=1, value=label)
        if bold:
            ws4.cell(row=r, column=1).font = BOLD_FONT
        if year0_formula is not None:
            _formula_cell(ws4, r, 2, year0_formula, fmt, bold=bold)
        for yi, ycol in enumerate(YEAR_COLS, start=1):
            month_num = (yi - 1) * 12 + 1 if take == "first" else yi * 12
            cl = _col(MONTH_COLS[month_num - 1])
            _formula_cell(ws4, r, ycol, f"='Monthly Cash Flow'!{cl}{month_row_ref}", fmt, bold=bold)
        this_row = r
        r += 1
        return this_row

    revenue_actual = annual["revenue"].actual_value if annual.get("revenue") else None
    expenses_actual = annual["expenses"].actual_value if annual.get("expenses") else None

    gpr_row = annual_flow_row("Gross Potential Rent (Total Cash Rent)", gpr_m_row)
    vac_credit_row = annual_flow_row("Less: Vacancy & Credit Loss", vac_credit_m_row)
    egi_row = annual_flow_row("Effective Rental Income", egi_m_row, bold=True)
    opex_row = annual_flow_row(
        "Operating Expenses", total_opex_m_row, year0_formula=f"=-{expenses_actual or 0}/{months_ytd}*12"
    )
    nonop_row = annual_flow_row(
        "Non-operating Expenses", nonop_m_row, year0_formula=f"=-{nonop_actual or 0}/{months_ytd}*12"
    )
    recovery_row = annual_flow_row("Recovery Income", recovery_m_row, year0_formula=f"=RecoveryPct*-B{opex_row}")
    total_rev_row = annual_flow_row(
        "Total Revenue", total_rev_m_row, bold=True, year0_formula=f"=B{egi_row}+B{recovery_row}"
    )
    noi_row = annual_flow_row(
        "NOI", noi_m_row, bold=True, year0_formula=f"=B{total_rev_row}+B{opex_row}+B{nonop_row}"
    )
    ti_lc_row = annual_flow_row("Less: TI / LC", less_ti_lc_m_row)
    capex_budget_row_val = f"=-{capex_budget or 0}"
    capex_row = annual_flow_row("Less: Capital Expenditures", capex_m_row, year0_formula=capex_budget_row_val)
    ws4.cell(row=capex_row, column=2).comment = Comment(
        "Year 0 uses the real full-year Capital Expenditures BUDGET figure (not YTD-annualized — "
        "capex isn't evenly spread through the year). Escalated by Capital Growth thereafter as a "
        "placeholder; replace with a real multi-year CapEx plan if/when one is dropped into the app.",
        "Deal Dashboard",
    )
    cfbds_row = annual_flow_row(
        "Cash Flow Before Debt Service", cfbds_m_row, bold=True,
        year0_formula=f"=B{noi_row}+B{ti_lc_row}+B{capex_row}",
    )

    r += 1
    ws4.cell(row=r, column=1, value="Debt Schedule (rolled up from Monthly Cash Flow)").font = SECTION_FONT
    r += 1

    beginning_row = annual_point_row(
        "Beginning Balance (all tranches)", total_beginning_balance_m_row, take="first",
    )
    interest_row = annual_flow_row("Interest Payment", total_interest_m_row)
    principal_row = annual_flow_row("Principal Payment", total_principal_m_row)
    ending_row = annual_point_row("Ending Balance (all tranches)", total_ending_balance_m_row, take="last")
    total_ds_row = annual_flow_row("Total Debt Service", total_ds_m_row, bold=True)
    # No Year 0 figure here — Year 0's debt service isn't modeled either (real full-year debt
    # service isn't cleanly derivable from the same partial-year YTD data the other Year 0
    # figures use), so a Year 0 Levered CF would misleadingly imply zero debt service.
    levered_cf_row = annual_flow_row("Levered Cash Flow (before sale)", levered_cf_m_row, bold=True)

    ws4.freeze_panes = "B5"

    # =======================================================================
    # Sheet 6: Returns
    # =======================================================================
    ws5 = out.create_sheet("Returns")
    ws5.sheet_view.showGridLines = False
    _autosize(ws5, [40] + [14] * (N_YEARS + 1))

    ws5["A1"] = "Returns"
    ws5["A1"].font = TITLE_FONT
    ws5["A2"] = (
        "Applies the Hold Period gate to the Cash Flow Projection sheet and adds net sale proceeds "
        "in the exit year. Exit uses the exit year's own (trailing) NOI x Cap Rate, not forward NOI."
    )
    ws5["A2"].font = NOTE_FONT
    ws5.merge_cells("A2:F2")

    r = 4
    _year_header_row(ws5, r)
    r += 2

    ws5.cell(row=r, column=1, value="Is Exit Year?")
    for i, col in enumerate(YEAR_COLS, start=1):
        _formula_cell(ws5, r, col, f"=IF({i}=HoldPeriod,1,0)", "0")
    exit_flag_row = r
    r += 1

    ws5.cell(row=r, column=1, value="Exit Year NOI (trailing)")
    for i, col in enumerate(YEAR_COLS, start=1):
        cl = _col(col)
        _formula_cell(ws5, r, col, f"=IF({cl}{exit_flag_row}=1,'Cash Flow Projection'!{cl}{noi_row},0)", MONEY_FMT)
    exit_noi_row = r
    r += 1

    ws5.cell(row=r, column=1, value="Gross Sale Value")
    for col in YEAR_COLS:
        cl = _col(col)
        _formula_cell(ws5, r, col, f"=IF({cl}{exit_flag_row}=1,{cl}{exit_noi_row}/CapRate,0)", MONEY_FMT)
    gross_sale_row = r
    r += 1

    ws5.cell(row=r, column=1, value="Less: Cost of Sale")
    for col in YEAR_COLS:
        cl = _col(col)
        _formula_cell(ws5, r, col, f"=-{cl}{gross_sale_row}*CostOfSale", MONEY_FMT)
    cost_sale_row = r
    r += 1

    ws5.cell(row=r, column=1, value="Less: Remaining Debt Balance at Exit")
    for col in YEAR_COLS:
        cl = _col(col)
        _formula_cell(ws5, r, col, f"=IF({cl}{exit_flag_row}=1,-'Cash Flow Projection'!{cl}{ending_row},0)", MONEY_FMT)
    payoff_row = r
    r += 1

    ws5.cell(row=r, column=1, value="Net Sale Proceeds").font = BOLD_FONT
    for col in YEAR_COLS:
        cl = _col(col)
        _formula_cell(ws5, r, col, f"={cl}{gross_sale_row}+{cl}{cost_sale_row}+{cl}{payoff_row}", MONEY_FMT, bold=True)
    net_sale_row = r
    r += 2

    ws5.cell(row=r, column=1, value="Cash Flow to Equity").font = BOLD_FONT
    _formula_cell(ws5, r, 2, "=-StartingEquity", MONEY_FMT, bold=True)
    for i, col in enumerate(YEAR_COLS, start=1):
        cl = _col(col)
        formula = f"=IF({i}<=HoldPeriod,'Cash Flow Projection'!{cl}{levered_cf_row}+{cl}{net_sale_row},0)"
        _formula_cell(ws5, r, col, formula, MONEY_FMT, bold=True)
    cf_to_equity_row = r
    r += 2

    ws5.cell(row=r, column=1, value="IRR")
    _formula_cell(ws5, r, 2, f"=IRR(B{cf_to_equity_row}:{_col(YEAR_COLS[-1])}{cf_to_equity_row})", PCT_FMT, bold=True)
    r += 1
    ws5.cell(row=r, column=1, value="Equity Multiple")
    _formula_cell(
        ws5, r, 2,
        f'=SUMIF(C{cf_to_equity_row}:{_col(YEAR_COLS[-1])}{cf_to_equity_row},">0")/-B{cf_to_equity_row}',
        MULT_FMT, bold=True,
    )

    ws5.freeze_panes = "B5"

    for name, ref in names.items():
        out.defined_names[name] = DefinedName(name, attr_text=ref)

    return out
