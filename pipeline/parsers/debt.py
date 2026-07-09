"""Parses a loan-tranche tab (e.g. "Rev Labs Interest") into a DebtSummary.

Layout observed in the Revolution Labs workbook has a specific trap: the rows
labeled "Interest Rate - Note A/B/Mezz" directly under "Outstanding Balance"
actually hold dollar principal balances, not rates. The real decimal rates live
under identically-worded labels several rows below, after a "SOFR*" row. This
parser anchors both blocks by position (N rows below "Outstanding Balance" /
"SOFR") rather than re-searching that ambiguous label text, so it can't
accidentally read a rate value as a balance or vice versa.
"""

from __future__ import annotations

from datetime import date
from typing import List, Optional

from openpyxl.worksheet.worksheet import Worksheet

from pipeline.models import DebtSummary, DebtTranche
from pipeline.parsers._shared import find_period_header_row, find_row_by_label, nearest_col_for_date


def parse_debt_tranches(
    ws: Worksheet, property_code: str, as_of: Optional[date] = None
) -> DebtSummary:
    header_row, date_cols = find_period_header_row(ws, min_hits=2)
    outstanding_row = find_row_by_label(ws, "Outstanding Balance")
    sofr_row = find_row_by_label(ws, "SOFR*") or find_row_by_label(ws, "SOFR", exact=False)
    stop_row = (
        find_row_by_label(ws, "Paydown")
        or find_row_by_label(ws, "Interest Carry")
        or sofr_row
    )

    if header_row is None or outstanding_row is None or sofr_row is None or stop_row is None:
        return DebtSummary(property_code=property_code, tranches=[])

    tranche_rows: List[int] = [
        r
        for r in range(outstanding_row + 1, stop_row)
        if isinstance(ws.cell(row=r, column=2).value, str) and ws.cell(row=r, column=2).value.strip()
    ]
    rate_rows = list(range(sofr_row + 1, sofr_row + 1 + len(tranche_rows)))

    target_col = nearest_col_for_date(ws, header_row, date_cols, as_of)
    as_of_date = ws.cell(row=header_row, column=target_col).value.date() if target_col else None

    tranches = []
    for trow, rrow in zip(tranche_rows, rate_rows):
        raw_name = str(ws.cell(row=trow, column=2).value or "")
        name = raw_name.replace("Interest Rate - ", "").strip() or raw_name.strip()
        balance = ws.cell(row=trow, column=target_col).value if target_col else None
        rate = ws.cell(row=rrow, column=target_col).value if target_col else None
        tranches.append(
            DebtTranche(
                tranche_name=name,
                outstanding_balance=float(balance) if isinstance(balance, (int, float)) else 0.0,
                interest_rate=float(rate) if isinstance(rate, (int, float)) else 0.0,
                as_of=as_of_date,
            )
        )

    sofr_val = ws.cell(row=sofr_row, column=target_col).value if target_col else None
    return DebtSummary(
        property_code=property_code,
        tranches=tranches,
        as_of=as_of_date,
        sofr_as_of=float(sofr_val) if isinstance(sofr_val, (int, float)) else None,
    )
