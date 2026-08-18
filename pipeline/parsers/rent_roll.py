"""Parses a Yardi "Tenancy Schedule II" rent roll export into a RentRollResult.

Layout observed (`TenancyScheduleII07_09_2026.xlsx` / `...08_18_2026.xlsx`,
sheet "Report1"): a multi-row-per-lease block. The FIRST row of each block
carries the property name in column 1 (the reliable anchor for "this is a
new lease") plus the core lease fields. Every other row in that block
(blank column 1) up to the next primary row carries one or both of: a
charge-type line ("Base Rent" or a "Recovery - ..." reimbursement — a lease
can have several recovery lines, e.g. RE-Tax + Operating) and/or a dated
future rent-escalation step — plus occasionally a second physical unit on a
multi-suite lease (e.g. an office suite plus a penthouse), whose own area is
already folded into the primary row's `lease_area`, so its row doesn't need
separate parsing.

**Column positions shift between export runs of this same report** —
confirmed directly: the 07_09_2026 export and the 08_18_2026 export differ
by one inserted "Desc" column, which shifts every column after it by one,
plus a block of new "Next Review"/increase-type columns appended at the
end. Column lookups below therefore go through `find_header_col()`
(`_shared.py`), which finds a column by its header TEXT rather than a fixed
position, so both the already-stored old-format files and any new export
keep parsing correctly without a version flag. The one exception is the
tenant-name column: its header text ("Lease"/"Customer") is confirmed
*not* to match where the real data actually sits (a known, stable quirk of
this report, unaffected by the other columns shifting), so that one column
stays a hardcoded position rather than a header lookup.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime
from typing import List, Optional

from openpyxl.worksheet.worksheet import Worksheet

from pipeline.models import ChargeLine, RentRollLine, RentRollResult, RentStep
from pipeline.parsers._shared import find_header_col

_AS_OF_RE = re.compile(r"As of Date:\s*(\d{1,2}/\d{1,2}/\d{4})")

_TENANT_NAME_COL = 7  # header text here is misleading — see module docstring


@dataclass
class _Columns:
    unit_area: Optional[int]
    lease_from: Optional[int]
    lease_to: Optional[int]
    term: Optional[int]
    lease_area: Optional[int]
    annual_rent: Optional[int]
    annual_rent_psf: Optional[int]
    lease_type: Optional[int]
    loc: Optional[int]
    charge_code: Optional[int]
    desc: Optional[int]
    charge_amt: Optional[int]
    step_date: Optional[int]
    step_annual: Optional[int]
    step_psf: Optional[int]


def _resolve_columns(ws: Worksheet) -> _Columns:
    charge_code = find_header_col(ws, "charge code") or find_header_col(ws, "charge type")
    step_date = find_header_col(ws, "start date")
    return _Columns(
        unit_area=find_header_col(ws, "unit area"),
        lease_from=find_header_col(ws, "lease from"),
        lease_to=find_header_col(ws, "lease to"),
        term=find_header_col(ws, "term"),
        lease_area=find_header_col(ws, "lease area"),
        annual_rent=find_header_col(ws, "annual rent"),
        annual_rent_psf=find_header_col(ws, "annual", "rent/area"),
        lease_type=find_header_col(ws, "lease type"),
        loc=find_header_col(ws, "loc amount") or find_header_col(ws, "bank guarantee"),
        charge_code=charge_code,
        desc=find_header_col(ws, "desc"),
        charge_amt=find_header_col(ws, "annual amt", after=charge_code) if charge_code else None,
        step_date=step_date,
        step_annual=find_header_col(ws, "rent step", "annual"),
        step_psf=find_header_col(ws, "annual", "rent/area", after=step_date) if step_date else None,
    )


def _cell(ws: Worksheet, r: int, col: Optional[int]):
    return ws.cell(row=r, column=col).value if col else None


def _parse_as_of(ws: Worksheet) -> Optional[date]:
    for row in ws.iter_rows(min_row=1, max_row=4, max_col=1):
        v = row[0].value
        if isinstance(v, str):
            match = _AS_OF_RE.search(v)
            if match:
                return datetime.strptime(match.group(1), "%m/%d/%Y").date()
    return None


def _to_date(v) -> Optional[date]:
    return v.date() if isinstance(v, datetime) else None


def _to_float(v) -> Optional[float]:
    return float(v) if isinstance(v, (int, float)) else None


def _is_primary_row(ws: Worksheet, r: int, cols: _Columns) -> bool:
    property_cell = ws.cell(row=r, column=1).value
    if not isinstance(property_cell, str) or not property_cell.strip():
        return False
    tenant_raw = ws.cell(row=r, column=_TENANT_NAME_COL).value
    tenant_name = tenant_raw.strip() if isinstance(tenant_raw, str) else ""
    is_vacant = tenant_name.upper() == "VACANT"
    # Column 1 is non-empty on the title/property-header/column-header rows
    # too (rows 1-3) — a real lease row always has either a lease-from date
    # or the literal "VACANT" marker, which those rows never do.
    lease_from_cell = _cell(ws, r, cols.lease_from)
    return is_vacant or isinstance(lease_from_cell, datetime)


def _parse_block_charges_and_steps(ws: Worksheet, start: int, end: int, cols: _Columns):
    """Scans rows [start, end) — a lease's primary row plus every row up to
    (not including) the next lease's primary row — for charge-type lines and
    rent-escalation steps. The two checks are independent per row (rows exist
    with both populated at once), so neither classification excludes the
    other."""
    charges: List[ChargeLine] = []
    rent_steps: List[RentStep] = []
    for r in range(start, end):
        charge_type = _cell(ws, r, cols.charge_code)
        charge_amt = _cell(ws, r, cols.charge_amt)
        if isinstance(charge_type, str) and charge_type.strip() and isinstance(charge_amt, (int, float)):
            desc_val = _cell(ws, r, cols.desc)
            charges.append(
                ChargeLine(
                    charge_type=charge_type.strip(),
                    annual_amount=float(charge_amt),
                    description=desc_val.strip() if isinstance(desc_val, str) else "",
                )
            )

        step_date = _cell(ws, r, cols.step_date)
        step_rent = _cell(ws, r, cols.step_annual)
        if isinstance(step_date, datetime) and isinstance(step_rent, (int, float)):
            rent_steps.append(
                RentStep(
                    effective_date=step_date.date(),
                    annual_rent=float(step_rent),
                    annual_rent_psf=_to_float(_cell(ws, r, cols.step_psf)),
                )
            )
    rent_steps.sort(key=lambda s: s.effective_date)
    return charges, rent_steps


def parse_rent_roll(ws: Worksheet, property_code: str) -> RentRollResult:
    as_of = _parse_as_of(ws)
    cols = _resolve_columns(ws)
    primary_rows = [r for r in range(1, ws.max_row + 1) if _is_primary_row(ws, r, cols)]

    lines = []
    for idx, r in enumerate(primary_rows):
        block_end = primary_rows[idx + 1] if idx + 1 < len(primary_rows) else ws.max_row + 1

        tenant_raw = ws.cell(row=r, column=_TENANT_NAME_COL).value
        tenant_name = tenant_raw.strip() if isinstance(tenant_raw, str) else ""
        is_vacant = tenant_name.upper() == "VACANT"

        charges, rent_steps = _parse_block_charges_and_steps(ws, r, block_end, cols)

        lines.append(
            RentRollLine(
                building=str(ws.cell(row=r, column=2).value or "").strip(),
                floor=str(ws.cell(row=r, column=3).value or "").strip(),
                unit_code=str(ws.cell(row=r, column=4).value or "").strip(),
                unit_area=_to_float(_cell(ws, r, cols.unit_area)),
                tenant_name="" if is_vacant else tenant_name,
                lease_from=_to_date(_cell(ws, r, cols.lease_from)),
                lease_to=_to_date(_cell(ws, r, cols.lease_to)),
                term_months=_to_float(_cell(ws, r, cols.term)),
                lease_area=_to_float(_cell(ws, r, cols.lease_area)),
                annual_rent=_to_float(_cell(ws, r, cols.annual_rent)),
                annual_rent_psf=_to_float(_cell(ws, r, cols.annual_rent_psf)),
                lease_type=str(_cell(ws, r, cols.lease_type) or "").strip(),
                is_vacant=is_vacant,
                loc_amount=_to_float(_cell(ws, r, cols.loc)),
                charges=charges,
                rent_steps=rent_steps,
            )
        )

    return RentRollResult(property_code=property_code, as_of=as_of, lines=lines)
