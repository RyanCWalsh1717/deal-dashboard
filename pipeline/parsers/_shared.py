"""Small helpers shared across parsers — mostly label/date-cell scanning, since
none of these workbook tabs have fixed row/column positions we can hardcode."""

from __future__ import annotations

import re
from datetime import date, datetime
from typing import List, Optional, Tuple

from openpyxl.worksheet.worksheet import Worksheet

_AS_OF_RE = re.compile(r"as of\s+([A-Za-z]+)\s+(\d{4})", re.IGNORECASE)


def parse_as_of_date(label: str) -> Optional[date]:
    """Parses labels like 'As of June 2026' into date(2026, 6, 1)."""
    match = _AS_OF_RE.search(label or "")
    if not match:
        return None
    try:
        return datetime.strptime(f"{match.group(1)} {match.group(2)}", "%B %Y").date()
    except ValueError:
        return None


def find_period_header_row(
    ws: Worksheet, max_row: int = 15, max_col: int = 60, min_hits: int = 2
) -> Tuple[Optional[int], List[int]]:
    """Finds the row holding actual `datetime` period cells and returns
    (row_index, [column indices with a datetime value]) or (None, [])."""
    for r in range(1, max_row + 1):
        cols = [c for c in range(1, max_col + 1) if isinstance(ws.cell(row=r, column=c).value, datetime)]
        if len(cols) >= min_hits:
            return r, cols
    return None, []


def find_row_by_label(
    ws: Worksheet, label: str, label_col: int = 2, max_row: int = 60, exact: bool = True
) -> Optional[int]:
    needle = label.strip().lower()
    for r in range(1, max_row + 1):
        v = ws.cell(row=r, column=label_col).value
        if not isinstance(v, str):
            continue
        candidate = v.strip().lower()
        if (exact and candidate == needle) or (not exact and needle in candidate):
            return r
    return None


def detect_code_label_cols(
    ws: Worksheet, start_row: int, max_scan: int = 15, max_col: int = 6, max_label_offset: int = 4
) -> Tuple[Optional[int], Optional[int]]:
    """Finds a (code_column, label_column) pair by scanning for a numeric code
    cell followed within a few columns by a text cell containing at least one
    letter (skips a plain numeric-code-repeated-as-a-string column, e.g. the
    Budget tab's column 2, which duplicates column 1's code as text)."""
    for r in range(start_row, start_row + max_scan):
        for code_c in range(1, max_col + 1):
            code_v = ws.cell(row=r, column=code_c).value
            if not (isinstance(code_v, (int, float)) and code_v != 0):
                continue
            for offset in range(1, max_label_offset + 1):
                label_c = code_c + offset
                label_v = ws.cell(row=r, column=label_c).value
                if isinstance(label_v, str) and label_v.strip() and any(ch.isalpha() for ch in label_v):
                    return code_c, label_c
    return None, None


def find_label_near(ws: Worksheet, row: int, base_col: int, max_extra: int = 2) -> Optional[str]:
    """Total/subtotal rows sometimes shift their label one or two columns right
    of where regular line-item labels sit (confirmed in the Budget tab: line
    items label at column 3, but "Total Base Rent"-style rows label at column
    4). Checks base_col..base_col+max_extra for the first non-empty string
    containing a letter, so callers don't have to hardcode which offset a
    given tab uses."""
    for c in range(base_col, base_col + max_extra + 1):
        v = ws.cell(row=row, column=c).value
        if isinstance(v, str) and v.strip() and any(ch.isalpha() for ch in v):
            return v.strip()
    return None


def extract_year(text: str) -> Optional[int]:
    match = re.search(r"(20\d{2})", text or "")
    return int(match.group(1)) if match else None


def nearest_col_for_date(
    ws: Worksheet, header_row: int, date_cols: List[int], target: Optional[date]
) -> Optional[int]:
    """Returns the column whose header date matches `target` exactly, else the
    latest column on/before `target`, else the last available column."""
    if not date_cols:
        return None
    if target is None:
        return date_cols[-1]

    dated = [(c, ws.cell(row=header_row, column=c).value.date()) for c in date_cols]
    for c, d in dated:
        if d == target:
            return c
    before = [c for c, d in dated if d <= target]
    return before[-1] if before else date_cols[-1]
