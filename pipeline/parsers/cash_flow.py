"""Parses a monthly GL cash-flow tab (e.g. "Cash Flow") into a CashFlowResult.

Layout observed in the Revolution Labs workbook: a multi-row header block (year
label / actual-vs-budget label / period-index label / actual date row), then
account rows with the GL code in one column and the label in the next.

Rows with no account code fall into two buckets: pure section headers (e.g.
"Base Rent") that carry no numeric values and are skipped entirely, and
label-only *rollup* rows (e.g. "Total Income", "Net Operating Income/(Loss)",
"Cash Flow after Debt Service") that DO carry monthly values — those are
captured into `subtotals` rather than discarded, since the workbook already
computes exactly the P&L rollups the Budget vs. Actuals summary view needs.

The header row and code/label columns are detected rather than hardcoded, since
column position isn't guaranteed to be identical across every property's workbook.
"""

from __future__ import annotations

from typing import Dict, List, Optional

from openpyxl.worksheet.worksheet import Worksheet

from pipeline.models import CashFlowLine, CashFlowResult
from pipeline.parsers._shared import detect_code_label_cols, find_label_near, find_period_header_row

_SCENARIO_WORDS = ("actual", "budget", "reforecast")


def _find_scenario_row(ws: Worksheet, header_row: int, date_cols: List[int], max_lookback: int = 6) -> Optional[int]:
    for r in range(max(1, header_row - max_lookback), header_row):
        matches = sum(
            1
            for c in date_cols
            if isinstance(ws.cell(row=r, column=c).value, str)
            and ws.cell(row=r, column=c).value.strip().lower() in _SCENARIO_WORDS
        )
        if matches >= 2:
            return r
    return None


def parse_cash_flow(ws: Worksheet, property_code: str) -> CashFlowResult:
    header_row, datetime_cols = find_period_header_row(ws, min_hits=3)
    if header_row is None:
        return CashFlowResult(property_code=property_code)

    code_col, label_col = detect_code_label_cols(ws, header_row + 1)
    if code_col is None:
        return CashFlowResult(property_code=property_code)

    col_to_period = {c: ws.cell(row=header_row, column=c).value.date() for c in datetime_cols}
    period_columns = sorted(set(col_to_period.values()))

    period_scenario: Dict = {}
    scenario_row = _find_scenario_row(ws, header_row, datetime_cols)
    if scenario_row is not None:
        for c, period in col_to_period.items():
            v = ws.cell(row=scenario_row, column=c).value
            if isinstance(v, str) and v.strip():
                period_scenario[period] = v.strip().lower()

    lines: List[CashFlowLine] = []
    subtotals: Dict[str, Dict] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        code_val = ws.cell(row=r, column=code_col).value
        has_code = isinstance(code_val, (int, float)) and code_val != 0

        if has_code:
            label_val = ws.cell(row=r, column=label_col).value
            if not isinstance(label_val, str) or not label_val.strip():
                continue
            label = label_val.strip()
        else:
            label = find_label_near(ws, r, label_col)
            if label is None:
                continue

        monthly_values = {}
        for c, period in col_to_period.items():
            val = ws.cell(row=r, column=c).value
            if isinstance(val, (int, float)):
                monthly_values[period] = float(val)

        if has_code:
            lines.append(
                CashFlowLine(
                    account_code=str(int(code_val)),
                    account_label=label,
                    monthly_values=monthly_values,
                )
            )
        elif monthly_values:
            subtotals[label] = monthly_values

    return CashFlowResult(
        property_code=property_code,
        period_columns=period_columns,
        lines=lines,
        subtotals=subtotals,
        period_scenario=period_scenario,
    )
