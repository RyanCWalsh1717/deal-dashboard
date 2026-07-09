"""Parses a "Distribution Recommendation" tab into a WaterfallTier — the
*actual* distribution waterfall (as opposed to a projected/promote waterfall,
which has no source data yet; see models.ProjectedDistributionWaterfall).

Layout observed: labels for the cash-projected/holdback/net-cash/recommendation
figures live in column C with values in column D. The investor table's header
row has "Ownership %" in some column; every other column of interest (display
name, legal entity, distribution amount, distribution %, contributions/
distributions to date, net capital after) is read at a fixed offset from that
column, not a hardcoded letter, since the header itself is offset one column
left of the data rows below it. The title row and the recommendation row both
contain the substring "Distribution Recommendation" — disambiguated by
requiring the matched row to also have a numeric value, which the title row
never does.
"""

from __future__ import annotations

from typing import Optional

from openpyxl.worksheet.worksheet import Worksheet

from pipeline.models import InvestorDistribution, WaterfallTier
from pipeline.parsers._shared import find_row_by_label


def _labeled_value(ws: Worksheet, needle: str, label_col: int = 3, value_col: int = 4, max_row: int = 20) -> Optional[float]:
    needle_l = needle.lower()
    for r in range(1, max_row + 1):
        label = ws.cell(row=r, column=label_col).value
        value = ws.cell(row=r, column=value_col).value
        if isinstance(label, str) and needle_l in label.lower() and isinstance(value, (int, float)):
            return float(value)
    return None


def _find_ownership_header(ws: Worksheet, max_row: int = 30, max_col: int = 15):
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip().lower() == "ownership %":
                return r, c
    return None, None


def _num(ws: Worksheet, row: int, col: int) -> Optional[float]:
    v = ws.cell(row=row, column=col).value
    return float(v) if isinstance(v, (int, float)) else None


def parse_distribution_waterfall(ws: Worksheet, tier_id: str, distributing_entity: str = "") -> WaterfallTier:
    as_of_row = find_row_by_label(ws, "As of", label_col=3, exact=False, max_row=6)
    as_of_label = str(ws.cell(row=as_of_row, column=3).value) if as_of_row else ""

    if not distributing_entity:
        title = ws.cell(row=1, column=3).value
        distributing_entity = str(title).strip() if isinstance(title, str) else ""

    cash_projected = _labeled_value(ws, "Cash Projected")
    net_cash_available = _labeled_value(ws, "Net Cash Available")
    distribution_recommendation = _labeled_value(ws, "Distribution Recommendation")

    holdback_start = find_row_by_label(ws, "Cash Hold-backs", label_col=3, exact=False, max_row=20)
    holdback_end = find_row_by_label(ws, "Total Cash Hold backs", label_col=3, exact=False, max_row=20)
    holdbacks = {}
    if holdback_start and holdback_end:
        for r in range(holdback_start + 1, holdback_end):
            label = ws.cell(row=r, column=3).value
            value = ws.cell(row=r, column=4).value
            if isinstance(label, str) and label.strip() and isinstance(value, (int, float)):
                holdbacks[label.strip()] = float(value)

    investors = []
    header_row, ownership_col = _find_ownership_header(ws)
    if header_row and ownership_col:
        r = header_row + 1
        while r < header_row + 30:
            legal = ws.cell(row=r, column=ownership_col - 1).value
            if not isinstance(legal, str) or not legal.strip():
                break
            display_raw = ws.cell(row=r, column=ownership_col - 2).value
            display_name = display_raw.strip() if isinstance(display_raw, str) and display_raw.strip() else legal.strip()

            investors.append(
                InvestorDistribution(
                    display_name=display_name,
                    legal_entity=legal.strip(),
                    ownership_pct=_num(ws, r, ownership_col) or 0.0,
                    distribution_amount=_num(ws, r, ownership_col + 3) or 0.0,
                    distribution_pct_of_contributions=_num(ws, r, ownership_col + 4),
                    contributions_to_date=_num(ws, r, ownership_col + 6),
                    distributions_to_date=_num(ws, r, ownership_col + 7),
                    net_capital_after=_num(ws, r, ownership_col + 8),
                )
            )
            r += 1

    return WaterfallTier(
        tier_id=tier_id,
        distributing_entity=distributing_entity,
        as_of_label=as_of_label,
        cash_projected=cash_projected,
        cash_holdbacks=holdbacks,
        net_cash_available=net_cash_available,
        distribution_recommendation=distribution_recommendation,
        investors=investors,
    )
