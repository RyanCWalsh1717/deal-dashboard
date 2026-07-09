"""Orchestrator: opens the distribution workbook once and calls every
sub-parser, resolving tab names through `PropertyConfig.sheet_map` so no
parser call ever hardcodes a tab name.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Union

import openpyxl

from pipeline.models import DistributionWaterfall, DistributionWorkbookResult
from pipeline.parsers._shared import parse_as_of_date
from pipeline.parsers.budget_vs_actual import build_pnl_summary, parse_budget_comparison, parse_budget_tab
from pipeline.parsers.cash_flow import parse_cash_flow
from pipeline.parsers.debt import parse_debt_tranches
from pipeline.parsers.equity import parse_equity_tab
from pipeline.parsers.waterfall import parse_distribution_waterfall
from pipeline.property_config import PropertyConfig


def parse_workbook(path: Union[str, Path], cfg: PropertyConfig) -> DistributionWorkbookResult:
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_map = cfg.sheet_map

    cash_flow = None
    cash_flow_sheet = sheet_map.get("cash_flow")
    if cash_flow_sheet and cash_flow_sheet in wb.sheetnames:
        cash_flow = parse_cash_flow(wb[cash_flow_sheet], cfg.property_code)

    equity = {}
    for logical_key in ("equity_lp", "equity_bhc"):
        sheet_name = sheet_map.get(logical_key)
        if sheet_name and sheet_name in wb.sheetnames:
            equity[logical_key] = parse_equity_tab(wb[sheet_name])

    # Match each ownership tier to a waterfall tab by tier shape (top-level tier
    # -> "waterfall_property" sheet; any nested tier -> "waterfall_cogp" sheet).
    # Good enough for the confirmed 1- and 2-tier cases; a 3rd tier would need a
    # richer per-tier sheet reference, deferred until a property needs one.
    tiers = {}
    as_of_date = None
    for tier in cfg.ownership_tiers:
        sheet_key = "waterfall_property" if tier.parent_tier is None else "waterfall_cogp"
        sheet_name = sheet_map.get(sheet_key)
        if not sheet_name or sheet_name not in wb.sheetnames:
            continue
        parsed_tier = parse_distribution_waterfall(wb[sheet_name], tier.tier_id, tier.distributing_entity)
        tiers[tier.tier_id] = parsed_tier
        if tier.parent_tier is None:
            as_of_date = parse_as_of_date(parsed_tier.as_of_label) or as_of_date

    waterfall = DistributionWaterfall(property_code=cfg.property_code, tiers=tiers) if tiers else None

    debt = None
    debt_sheet = sheet_map.get("debt")
    if debt_sheet and debt_sheet in wb.sheetnames:
        debt = parse_debt_tranches(wb[debt_sheet], cfg.property_code, as_of=as_of_date)

    budget_comparison = None
    budget_summary = None
    budget_sheet = sheet_map.get("budget_current")
    if budget_sheet and budget_sheet in wb.sheetnames and cash_flow is not None:
        budget_result = parse_budget_tab(wb[budget_sheet], cfg.property_code)
        # The real "today" for P&L data is the last period the Cash Flow tab
        # itself labels "Actuals" — NOT the distribution's as-of date, which
        # is a different concept (when the distribution was calculated) and
        # can land on a column the tab labels "Budget" (the current year's
        # forecast/reforecast months), producing a comparison that's silently
        # actual-vs-actual's-own-forecast instead of actual-vs-budget.
        period = cash_flow.last_actual_period() or as_of_date or cash_flow.latest_period()
        if period:
            budget_comparison = parse_budget_comparison(budget_result, cash_flow, period, cfg.property_code)
            budget_summary = build_pnl_summary(cash_flow, budget_result, period, cfg.property_code)

    return DistributionWorkbookResult(
        property_code=cfg.property_code,
        source_path=str(path),
        parsed_at=datetime.now(),
        cash_flow=cash_flow,
        equity=equity,
        debt=debt,
        waterfall=waterfall,
        projected_waterfall=None,  # no parser exists until the investment model is finalized
        budget_comparison=budget_comparison,
        budget_summary=budget_summary,
    )
