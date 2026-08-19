"""Parses a standalone Yardi "Budget Comparison" GL report (single sheet,
title cell literally reading "Budget Comparison") into a BudgetComparisonResult.

This is a full chart-of-accounts dump (Income through Equity, every GL code)
with a "Period = <Mon> <Year>" label and, per account, PTD Actual / PTD
Budget / Variance / % Var / YTD Actual / YTD Budget / Variance / % Var /
Annual columns (row 5 header) — richer than the distribution workbook's own
internal Budget tab (see budget_vs_actual.py), which only carries a handful
of P&L subtotals. Only the PTD Actual/PTD Budget columns are used here, to
match the account-level "Detailed" view's existing single-period-comparison
shape; YTD and Annual aren't consumed.

Leaf GL lines have a numeric code in column 1 and real values in every
numeric column. Section-header rows (e.g. "400000 INCOME") have no values at
all; subtotal/rollup rows (e.g. "419999 TOTAL BASE RENT") do have values but
their label always starts with "TOTAL" (real leaf labels are Title Case, e.g.
"Rent - Lab") — confirmed directly against Budget_Comparison_Accrual (31).xlsx.
"""

from __future__ import annotations

import re
from datetime import date, datetime
from typing import Dict, Optional

from openpyxl.worksheet.worksheet import Worksheet

from pipeline.models import BudgetComparisonResult, BudgetLine

_TITLE_ROW = 2
_PERIOD_ROW = 3
_FIRST_DATA_ROW = 6
_CODE_COL = 1
_LABEL_COL = 2
_PTD_ACTUAL_COL = 3
_PTD_BUDGET_COL = 4
_YTD_ACTUAL_COL = 7
_YTD_BUDGET_COL = 8

# Public aliases — for callers outside this module (app.py's BOMA-rollup
# wiring) that need to pick which column parse_opex_categories() reads.
PTD_ACTUAL_COL = _PTD_ACTUAL_COL
PTD_BUDGET_COL = _PTD_BUDGET_COL
YTD_ACTUAL_COL = _YTD_ACTUAL_COL
YTD_BUDGET_COL = _YTD_BUDGET_COL

_PERIOD_RE = re.compile(r"Period\s*=\s*([A-Za-z]+\s+\d{4})")


def is_budget_comparison_report(ws: Worksheet) -> bool:
    title = ws.cell(row=_TITLE_ROW, column=1).value
    return isinstance(title, str) and title.strip().lower() == "budget comparison"


def parse_period(ws: Worksheet) -> Optional[date]:
    cell = ws.cell(row=_PERIOD_ROW, column=1).value
    if not isinstance(cell, str):
        return None
    match = _PERIOD_RE.search(cell)
    if not match:
        return None
    text = match.group(1)
    for fmt in ("%B %Y", "%b %Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


_ANNUAL_COL = 11

# Maps our internal Annual Budget row keys to this report's own subtotal
# labels (column 2) — these are the rows parse_budget_comparison_report()
# deliberately excludes (label starts with "TOTAL"/is a rollup), but they're
# exactly the P&L waterfall the Annual Budget section needs. Confirmed
# directly against Budget_Comparison_Accrual (31).xlsx: this report's own
# "NET INCOME" row mathematically equals NOI minus Debt Service — i.e. our
# model's "Cash Flow after Debt Service", a different concept from the
# distribution workbook's trailing-12-month Net Income (already excluded
# from this section for that reason) — nothing labeled "Net Income" is ever
# shown to the user from this mapping.
_ANNUAL_ROW_LABELS = {
    "revenue": "TOTAL REVENUE",
    "expenses": "TOTAL EXPENSES",
    "non_operating_expenses": "TOTAL OPERATING EXPENSES - NON RECOVERABLE",
    "noi": "NET OPERATING INCOME",
    "debt_service": "TOTAL INTEREST EXPENSE & FEES",
    "cash_flow_after_debt_service": "NET INCOME",
}
# Capital Expenditures has no row of its own — it's the balance-sheet asset
# movement for the year, sign-flipped (an asset increase is a debit/negative
# in this report's convention). "CASH FLOW" is the bottom line, in column 2
# on a row with a blank column 1 (same as the other totals here).
_CAPEX_LABEL = "TOTAL ASSETS"
_CASH_FLOW_AFTER_CAPEX_LABEL = "CASH FLOW"


def _find_label_value(ws: Worksheet, label: str, label_col: int = _LABEL_COL, value_col: int = _ANNUAL_COL) -> Optional[float]:
    needle = label.strip().upper()
    for r in range(1, ws.max_row + 1):
        cell = ws.cell(row=r, column=label_col).value
        if isinstance(cell, str) and cell.strip().upper() == needle:
            value = ws.cell(row=r, column=value_col).value
            if isinstance(value, (int, float)):
                return float(value)
    return None


def parse_annual_budget_totals(ws: Worksheet) -> dict:
    """Reads the Annual-column value off each labeled subtotal row in
    _ANNUAL_ROW_LABELS, plus Capital Expenditures and Cash Flow after
    Capital Expenditures. Returns whatever subset of the 8 keys it finds —
    callers decide what to do with a partial result."""
    totals = {}
    for key, label in _ANNUAL_ROW_LABELS.items():
        value = _find_label_value(ws, label)
        if value is not None:
            totals[key] = value

    capex = _find_label_value(ws, _CAPEX_LABEL)
    if capex is not None:
        totals["capital_expenditures"] = -capex

    cash_flow_after_capex = _find_label_value(ws, _CASH_FLOW_AFTER_CAPEX_LABEL)
    if cash_flow_after_capex is not None:
        totals["cash_flow_after_capital_expenditures"] = cash_flow_after_capex

    return totals


def parse_ytd_annual_budget_totals(ws: Worksheet) -> dict:
    """Same 8-row P&L waterfall as parse_annual_budget_totals(), but reading
    the report's own YTD Actual/YTD Budget columns instead of PTD — this is
    the live reforecast the distribution model itself carries (YTD actual
    through the current period, budget for the full year as most recently
    reforecast), distinct from the Annual column's once-set original budget."""
    totals = {}
    for key, label in _ANNUAL_ROW_LABELS.items():
        actual = _find_label_value(ws, label, value_col=_YTD_ACTUAL_COL)
        budget = _find_label_value(ws, label, value_col=_YTD_BUDGET_COL)
        if actual is not None or budget is not None:
            totals[key] = (actual, budget)

    capex_actual = _find_label_value(ws, _CAPEX_LABEL, value_col=_YTD_ACTUAL_COL)
    capex_budget = _find_label_value(ws, _CAPEX_LABEL, value_col=_YTD_BUDGET_COL)
    if capex_actual is not None or capex_budget is not None:
        totals["capital_expenditures"] = (
            -capex_actual if capex_actual is not None else None,
            -capex_budget if capex_budget is not None else None,
        )

    cf_actual = _find_label_value(ws, _CASH_FLOW_AFTER_CAPEX_LABEL, value_col=_YTD_ACTUAL_COL)
    cf_budget = _find_label_value(ws, _CASH_FLOW_AFTER_CAPEX_LABEL, value_col=_YTD_BUDGET_COL)
    if cf_actual is not None or cf_budget is not None:
        totals["cash_flow_after_capital_expenditures"] = (cf_actual, cf_budget)

    return totals


# The report's own recoverable-OpEx subtotal rows — confirmed these 13 sum to
# exactly the "TOTAL OPERATING EXPENSES - RECOVERABLE" figure ($6,327,494.49
# against Budget_Comparison_Accrual (31).xlsx), so they fully account for it
# with no leftover/uncategorized bucket. Non-recoverable OpEx (~$43K) stays a
# single line elsewhere — small, and not asked to be broken out further.
OPEX_CATEGORY_LABELS = {
    "cleaning_janitorial": "TOTAL CLEANING/JANITORIAL",
    "utilities": "TOTAL UTILITIES",
    "general_repairs_maintenance": "TOTAL GENERAL REPAIRS & MAINTENANCE",
    "hvac_maintenance": "TOTAL HVAC MAINTENANCE",
    "plumbing": "TOTAL PLUMBING",
    "electrical_maintenance": "TOTAL ELECTRICAL MAINTENANCE",
    "security_fire_life_safety": "TOTAL SECURITY / FIRE / LIFE SAFETY",
    "elevator_maintenance": "TOTAL ELEVATOR MAINTENANCE",
    "landscaping": "TOTAL LANDSCAPING",
    "parking_garage_maintenance": "TOTAL PARKING AND GARAGE MAINTENANCE",
    "administrative": "TOTAL ADMINISTRATIVE",
    "insurance": "TOTAL INSURANCE",
    "real_estate_taxes": "TOTAL REAL ESTATE TAXES",
}


def parse_opex_categories(ws: Worksheet, value_col: int = _ANNUAL_COL) -> dict:
    """Reads the given column's value off each recoverable-OpEx category
    subtotal row (see OPEX_CATEGORY_LABELS) — defaults to the Annual column
    (the Hold/Sell model's use case), but pass _PTD_ACTUAL_COL/_PTD_BUDGET_COL
    or _YTD_ACTUAL_COL/_YTD_BUDGET_COL for the Budget vs. Actuals view's BOMA
    rollup. Returns whatever subset is found, keyed by our internal category
    name — callers decide what to do with a partial result (same convention
    as parse_annual_budget_totals())."""
    categories = {}
    for key, label in OPEX_CATEGORY_LABELS.items():
        value = _find_label_value(ws, label, value_col=value_col)
        if value is not None:
            categories[key] = value
    return categories


# Yardi chart-of-accounts reference for this property's real accounts (revlabpm) —
# every recoverable-OpEx category this report exposes, mapped to a BOMA-standard
# expense category, so the Budget vs. Actuals Summary view can show "Expenses" as
# a real breakdown instead of one lump number. This mapping is the thing to edit
# if Ryan wants categories grouped differently — same idea as ga-automation's own
# per-property GL account-role config, just for this report's category rollups
# rather than individual account codes (which the report itself doesn't expose at
# the leaf level for these totals — only the category subtotal, see the module
# docstring). Confirmed: the 13 categories below sum to exactly this report's own
# "TOTAL OPERATING EXPENSES - RECOVERABLE" figure — nothing left uncategorized.
BOMA_CATEGORY_MAP = {
    "cleaning_janitorial": "Cleaning",
    "utilities": "Utilities",
    "general_repairs_maintenance": "Repairs & Maintenance",
    "hvac_maintenance": "Repairs & Maintenance",
    "plumbing": "Repairs & Maintenance",
    "electrical_maintenance": "Repairs & Maintenance",
    "elevator_maintenance": "Repairs & Maintenance",
    "security_fire_life_safety": "Security",
    "landscaping": "Roads & Grounds",
    "parking_garage_maintenance": "Roads & Grounds",
    "administrative": "Administrative",
    "insurance": "Insurance",
    "real_estate_taxes": "Real Estate Taxes",
}

# Display order for the BOMA rollup table — not alphabetical, matches the order
# BOMA's own standard expense report typically lists these in.
BOMA_CATEGORY_ORDER = [
    "Cleaning",
    "Repairs & Maintenance",
    "Utilities",
    "Security",
    "Roads & Grounds",
    "Administrative",
    "Insurance",
    "Real Estate Taxes",
]


def boma_rollup(categories: dict) -> dict:
    """Sums a {internal_category_key: value} dict (from parse_opex_categories())
    into BOMA-standard buckets via BOMA_CATEGORY_MAP. Skips any category key not
    in the map rather than silently dropping it into the wrong bucket — that's a
    sign the map needs a new entry, not a value to lose."""
    rolled: Dict[str, float] = {}
    for key, value in categories.items():
        boma_key = BOMA_CATEGORY_MAP.get(key)
        if boma_key is None:
            continue
        rolled[boma_key] = rolled.get(boma_key, 0.0) + value
    return rolled


def parse_budget_comparison_report(
    ws: Worksheet, property_code: str, actual_col: int = _PTD_ACTUAL_COL, budget_col: int = _PTD_BUDGET_COL
) -> BudgetComparisonResult:
    """Leaf-level (account-code) actual/budget comparison. Defaults to the
    PTD columns (the Detailed sub-view's original shape); pass
    _YTD_ACTUAL_COL/_YTD_BUDGET_COL for a year-to-date leaf-level view."""
    period = parse_period(ws)
    lines = []
    for r in range(_FIRST_DATA_ROW, ws.max_row + 1):
        code = ws.cell(row=r, column=_CODE_COL).value
        label = ws.cell(row=r, column=_LABEL_COL).value
        actual = ws.cell(row=r, column=actual_col).value
        budget = ws.cell(row=r, column=budget_col).value

        if not isinstance(code, str) or not code.strip().isdigit():
            continue
        if not isinstance(label, str) or not label.strip() or label.strip().upper().startswith("TOTAL"):
            continue
        if not isinstance(actual, (int, float)) or not isinstance(budget, (int, float)):
            continue

        lines.append(
            BudgetLine(
                account_code=code.strip(),
                account_label=label.strip(),
                budget_value=float(budget),
                actual_value=float(actual),
            )
        )

    return BudgetComparisonResult(property_code=property_code, period=period, lines=lines)


def parse_ytd_budget_comparison_report(ws: Worksheet, property_code: str) -> BudgetComparisonResult:
    return parse_budget_comparison_report(ws, property_code, actual_col=_YTD_ACTUAL_COL, budget_col=_YTD_BUDGET_COL)
