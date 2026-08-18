"""Extraction of cash-account balances (escrows, reserves, operating cash) and
loan-statement data from two real, confirmed formats (validated 2026-07-09
against actual Revolution Labs documents):

1. Berkadia loan servicer statements (PDF) — one per tranche, e.g. "Revolution
   Labs - Note A1". Fixed two-column layout; only the left column (balance
   info) is parsed. Confirmed fields: Principal Balance, Interest Rate (the
   real all-in rate — NOT the same as the "rate" derived in debt.py from the
   distribution workbook, which is actually just the spread over SOFR), Tax
   Escrow Balance, Insurance Escrow Balance, Reserve Balance.
2. A Yardi trial balance export (.xlsx) — account code / label / Forward /
   Debit / Credit / Ending Balance columns, with an entity header row
   ("Property = <code> ...") marking the start of that entity's account
   block. A single export can cover more than one property, so this is
   entity-block-aware rather than a blind whole-file scan — extraction is
   filtered to the requested property_code's block only.
"""

from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Union

from pipeline.models import CashAccountBalance, EntityTrialBalance, LoanStatement

_HEADER_RE = re.compile(
    r"Property:\s*(?P<label>.+?)\s+Loan No:\s*(?P<loan_no>\S+)\s+Interest Rate:\s*(?P<rate>[\d.]+)"
)
_AS_OF_RE = re.compile(r"AS OF\s+(\d{2}/\d{2}/\d{4})")


def _money(text: str, label: str) -> Optional[float]:
    match = re.search(re.escape(label) + r"\s+\$?\s*([\d,]+\.\d{2})", text)
    return float(match.group(1).replace(",", "")) if match else None


def parse_loan_statement(path: Union[str, Path]) -> Optional[LoanStatement]:
    import pdfplumber

    with pdfplumber.open(path) as pdf:
        text = "\n".join(page.extract_text() or "" for page in pdf.pages)

    header = _HEADER_RE.search(text)
    if not header:
        return None

    property_label = header.group("label").strip()
    tranche_name = property_label.split(" - ")[-1].strip() if " - " in property_label else property_label

    as_of_match = _AS_OF_RE.search(text)
    as_of = datetime.strptime(as_of_match.group(1), "%m/%d/%Y").date() if as_of_match else None

    return LoanStatement(
        tranche_name=tranche_name,
        loan_number=header.group("loan_no"),
        interest_rate=float(header.group("rate")) / 100.0,
        as_of=as_of,
        principal_balance=_money(text, "Principal Balance"),
        interest_paid_ytd=_money(text, "Interest Paid YTD"),
        tax_escrow_balance=_money(text, "Tax Escrow Balance"),
        insurance_escrow_balance=_money(text, "Insurance Escrow Balance"),
        reserve_balance=_money(text, "Reserve Balance"),
        total_payment_due=_money(text, "Total Payment Due"),
    )


def loan_statement_cash_accounts(stmt: LoanStatement) -> List[CashAccountBalance]:
    """The escrow/reserve lines of a loan statement, as Cash-tab boxes."""
    boxes = []
    for label, value in (
        (f"Tax Escrow ({stmt.tranche_name})", stmt.tax_escrow_balance),
        (f"Insurance Escrow ({stmt.tranche_name})", stmt.insurance_escrow_balance),
        (f"Reserve ({stmt.tranche_name})", stmt.reserve_balance),
    ):
        if value is not None:
            boxes.append(
                CashAccountBalance(label=label, balance=value, source="loan_statement", as_of=stmt.as_of)
            )
    return boxes


_ENTITY_HEADER_RE = re.compile(r"Property\s*=\s*(\S+)")
_CASH_LABEL_KEYWORDS = ("cash - operating", "cash - development", "restricted cash", "escrow", "reserve")

# Balance-sheet accounts the Reconciliation section cross-checks against
# other source files. Matched by label keyword (like _CASH_LABEL_KEYWORDS
# above) rather than a hardcoded account code, so this doesn't break if a
# different property's chart of accounts numbers things differently.
_AP_LABEL_KEYWORDS = ("accounts payable",)
_AR_LABEL_KEYWORDS = ("accounts receivable - control", "accounts receivable - tenant billback")
_CONTRIBUTIONS_LABEL_KEYWORDS = ("contributions",)
_DISTRIBUTIONS_LABEL_KEYWORDS = ("distributions",)
_RETAINED_EARNINGS_LABEL_KEYWORDS = ("retained earnings",)
_MORTGAGE_LABEL_KEYWORDS = ("mortgage payable",)


def _parse_tb_as_of(ws) -> Optional[object]:
    for row in ws.iter_rows(min_row=1, max_row=6, max_col=1):
        v = row[0].value
        if isinstance(v, str) and v.strip().lower().startswith("period"):
            period_text = v.split("=", 1)[1].strip() if "=" in v else v.strip()
            try:
                return datetime.strptime(period_text, "%B %Y").date()
            except ValueError:
                pass
    return None


def _accumulate(current: Optional[float], label_lower: str, keywords: tuple, amount: float, positive: bool) -> Optional[float]:
    if not any(kw in label_lower for kw in keywords):
        return current
    signed = abs(amount) if positive else amount
    return (current or 0.0) + signed


def parse_entity_trial_balance(
    path: Union[str, Path], yardi_codes: Optional[List[str]] = None
) -> List[EntityTrialBalance]:
    """Parses every entity block in a Yardi trial balance export into a full
    EntityTrialBalance — cash/escrow accounts (same rows
    parse_trial_balance_cash_accounts() has always returned) plus the
    Accounts Receivable / Accounts Payable / Contributions / Distributions /
    Retained Earnings / Mortgage Payable figures the Reconciliation section
    needs. Restricted to entity block(s) matching `yardi_codes`, same
    convention as parse_trial_balance_cash_accounts() (see that docstring
    for why yardi_codes rather than the dashboard's own property_code)."""
    import openpyxl

    yardi_codes = yardi_codes or []
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    as_of = _parse_tb_as_of(ws)

    # Column layout: 1=code, 2=label, 3=forward, 4=debit, 5=credit, 6=ending balance.
    entities: List[EntityTrialBalance] = []
    current: Optional[EntityTrialBalance] = None
    in_target_block = not yardi_codes

    for r in range(1, ws.max_row + 1):
        col1 = ws.cell(row=r, column=1).value
        if isinstance(col1, str) and col1.strip().lower().startswith("property"):
            entity_match = _ENTITY_HEADER_RE.search(col1)
            block_code = entity_match.group(1) if entity_match else col1.strip()
            entity_name = col1.split("=", 1)[1].strip() if "=" in col1 else col1.strip()
            in_target_block = not yardi_codes or any(code in block_code for code in yardi_codes)
            current = EntityTrialBalance(entity_code=block_code, entity_name=entity_name, as_of=as_of) if in_target_block else None
            if current:
                entities.append(current)
            continue

        if not in_target_block or current is None:
            continue

        label = ws.cell(row=r, column=2).value
        ending = ws.cell(row=r, column=6).value
        if not isinstance(label, str) or not isinstance(ending, (int, float)):
            continue
        label_lower = label.strip().lower()
        amount = float(ending)

        if any(kw in label_lower for kw in _CASH_LABEL_KEYWORDS):
            current.cash_accounts.append(
                CashAccountBalance(
                    label=label.strip(),
                    balance=amount,
                    account_code=str(col1) if col1 is not None else "",
                    source="trial_balance",
                    as_of=as_of,
                    entity_code=current.entity_code,
                )
            )

        current.accounts_receivable = _accumulate(current.accounts_receivable, label_lower, _AR_LABEL_KEYWORDS, amount, positive=True)
        current.accounts_payable = _accumulate(current.accounts_payable, label_lower, _AP_LABEL_KEYWORDS, amount, positive=True)
        current.contributions = _accumulate(current.contributions, label_lower, _CONTRIBUTIONS_LABEL_KEYWORDS, amount, positive=True)
        current.distributions = _accumulate(current.distributions, label_lower, _DISTRIBUTIONS_LABEL_KEYWORDS, amount, positive=False)
        current.retained_earnings = _accumulate(current.retained_earnings, label_lower, _RETAINED_EARNINGS_LABEL_KEYWORDS, amount, positive=False)
        current.mortgage_payable = _accumulate(current.mortgage_payable, label_lower, _MORTGAGE_LABEL_KEYWORDS, amount, positive=True)

    return entities


def parse_trial_balance_cash_accounts(
    path: Union[str, Path], yardi_codes: Optional[List[str]] = None
) -> List[CashAccountBalance]:
    """Thin wrapper over parse_entity_trial_balance() for callers that only
    need the Cash-tab account boxes, not the full entity balance-sheet
    figures. If `yardi_codes` is empty/None, or no matching block is found,
    parse_entity_trial_balance() falls back to scanning the whole sheet
    (better than silently returning nothing, but callers should prefer
    passing yardi_codes)."""
    entities = parse_entity_trial_balance(path, yardi_codes)
    return [acct for entity in entities for acct in entity.cash_accounts]
