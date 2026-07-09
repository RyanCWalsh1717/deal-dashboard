"""Parses a Budget tab (same account-code/label shape as Cash Flow, but with
string "M1".."M12" period headers instead of real dates) and joins it against
an already-parsed CashFlowResult to build a BudgetComparisonResult.

Accounts present on only one side (budget or actual) get a None on the missing
side rather than being dropped, so gaps are visible in the UI instead of hidden.
"""

from __future__ import annotations

import re
from datetime import date
from typing import Dict, Optional

from openpyxl.worksheet.worksheet import Worksheet

from pipeline.models import BudgetComparisonResult, BudgetLine, CashFlowLine, CashFlowResult
from pipeline.parsers._shared import detect_code_label_cols, extract_year, find_label_near

_MONTH_RE = re.compile(r"^M(\d{1,2})$", re.IGNORECASE)


def _find_month_header_row(ws: Worksheet, max_row: int = 10, max_col: int = 40):
    for r in range(1, max_row + 1):
        cols = {}
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str):
                m = _MONTH_RE.match(v.strip())
                if m:
                    cols[c] = int(m.group(1))
        if cols:
            return r, cols
    return None, {}


def parse_budget_tab(ws: Worksheet, property_code: str, year: Optional[int] = None) -> CashFlowResult:
    """Returns a CashFlowResult (same shape as the actuals side) so the two can
    be compared with plain dict joins by account_code/period."""
    year = year or extract_year(ws.title)
    header_row, month_cols = _find_month_header_row(ws)
    if header_row is None or not year:
        return CashFlowResult(property_code=property_code)

    code_col, label_col = detect_code_label_cols(ws, header_row + 1)
    if code_col is None:
        return CashFlowResult(property_code=property_code)

    col_to_period = {c: date(year, month, 1) for c, month in month_cols.items() if 1 <= month <= 12}
    period_columns = sorted(set(col_to_period.values()))

    lines = []
    subtotals: Dict[str, Dict] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        code_val = ws.cell(row=r, column=code_col).value
        has_code = isinstance(code_val, (int, float)) and code_val != 0

        if has_code:
            label_val = ws.cell(row=r, column=label_col).value
            if not isinstance(label_val, str) or not label_val.strip():
                continue
            label = label_val.strip()
        else:
            label = find_label_near(ws, r, label_col)
            if label is None:
                continue

        monthly_values = {}
        for c, period in col_to_period.items():
            val = ws.cell(row=r, column=c).value
            if isinstance(val, (int, float)):
                monthly_values[period] = float(val)

        if has_code:
            lines.append(
                CashFlowLine(
                    account_code=str(int(code_val)),
                    account_label=label,
                    monthly_values=monthly_values,
                )
            )
        elif monthly_values:
            subtotals[label] = monthly_values

    return CashFlowResult(
        property_code=property_code, period_columns=period_columns, lines=lines, subtotals=subtotals
    )


def _sum_optional(*values: Optional[float]) -> Optional[float]:
    present = [v for v in values if v is not None]
    return sum(present) if present else None


def _sub_optional(*values: Optional[float]) -> Optional[float]:
    """First value minus every subsequent one; None if the first value is
    missing (a total can't be computed without its starting point)."""
    if not values or values[0] is None:
        return None
    total = values[0]
    for v in values[1:]:
        if v is None:
            return None
        total -= v
    return total


# Row label -> (actual subtotal label, budget subtotal label). Where the two
# tabs use different wording for the same concept (confirmed by inspecting
# both tabs directly — e.g. actual's "Total Interest Expense & Fees" vs
# budget's "Total Debt Service"), both labels are recorded here so a caller
# never has to guess which one applies to which side.
_DIRECT_SUBTOTAL_MAP = {
    "Revenue": ("Total Income", "Total Income"),
    "Non-operating Expenses": (
        "Total Operating Expenses - Unrecoverable",
        "Total Operating Expenses - Unrecoverable",
    ),
    "Debt Service": ("Total Interest Expense & Fees", "Total Debt Service"),
}


def build_pnl_summary(
    actual: CashFlowResult, budget: CashFlowResult, period: date, property_code: str
) -> BudgetComparisonResult:
    """Builds the 7-row P&L summary (Revenue / Expenses / Non-operating
    Expenses / NOI / Debt Service / Cash flow after debt and capital / Net
    Income) Ryan asked for.

    Only "Revenue", "Non-operating Expenses", and "Debt Service" have a
    directly matching subtotal label on both the actual (Cash Flow tab) and
    budget tab — see _DIRECT_SUBTOTAL_MAP. Everything else is either only
    present on one side (e.g. budget has no NOI or "cash flow after" rollups
    at all) or the two tabs use a different expense taxonomy (actual: one
    lump "Total Operating Expenses"; budget: split into "...Recoverable" and
    "...Non-Recoverable"). Those are computed here using the SAME formula
    order the Cash Flow tab itself uses (NOI = Income - OpEx; Cash Flow after
    Capital = NOI - Non-op - Debt Service - CapEx) rather than guessed at, but
    still surfaced as "budget" in the Missing column so Ryan can sanity-check
    the derivation against how he'd expect it categorized.
    """
    revenue_a = actual.subtotal("Total Income", period)
    revenue_b = budget.subtotal("Total Income", period)

    expenses_a = actual.subtotal("Total Operating Expenses", period)
    expenses_b = _sum_optional(
        budget.subtotal("Total Operating Expenses - Recoverable", period),
        budget.subtotal("Total Operating Expenses - Non-Recoverable", period),
    )

    nonop_a = actual.subtotal("Total Operating Expenses - Unrecoverable", period)
    nonop_b = budget.subtotal("Total Operating Expenses - Unrecoverable", period)

    noi_a = actual.subtotal("Net Operating Income/(Loss)", period)
    noi_b = _sub_optional(revenue_b, expenses_b)

    debt_a = actual.subtotal("Total Interest Expense & Fees", period)
    debt_b = budget.subtotal("Total Debt Service", period)

    capex_b = budget.subtotal("Total Capital Expenditures", period)
    cf_after_a = actual.subtotal("Cash Flow after Capital Expenditures", period)
    cf_after_b = _sub_optional(noi_b, nonop_b, debt_b, capex_b)

    # Confirmed 2026-07-09: the Cash Flow tab's "Net Income" row is a
    # trailing-12-month figure that only extends through the last actual
    # period (it has no forward/budget-scenario values at all) — so there's
    # no budget-side equivalent to compare against. Actual comes straight
    # from the distribution file; budget stays None rather than fabricated.
    net_income_a = actual.subtotal("Net Income", period)

    rows = [
        ("revenue", "Revenue", revenue_b, revenue_a),
        ("expenses", "Expenses", expenses_b, expenses_a),
        ("non_operating_expenses", "Non-operating Expenses", nonop_b, nonop_a),
        ("noi", "NOI", noi_b, noi_a),
        ("debt_service", "Debt Service", debt_b, debt_a),
        ("cash_flow_after_debt_and_capital", "Cash Flow after Debt and Capital", cf_after_b, cf_after_a),
        ("net_income", "Net Income", None, net_income_a),
    ]
    lines = [
        BudgetLine(account_code=code, account_label=label, budget_value=b, actual_value=a)
        for code, label, b, a in rows
    ]
    return BudgetComparisonResult(property_code=property_code, period=period, lines=lines)


def parse_budget_comparison(
    budget_result: CashFlowResult, actual_result: CashFlowResult, period: date, property_code: str
) -> BudgetComparisonResult:
    budget_by_code: Dict[str, CashFlowLine] = {line.account_code: line for line in budget_result.lines}
    actual_by_code: Dict[str, CashFlowLine] = {line.account_code: line for line in actual_result.lines}

    lines = []
    for code in sorted(set(budget_by_code) | set(actual_by_code)):
        budget_line = budget_by_code.get(code)
        actual_line = actual_by_code.get(code)
        label = (budget_line or actual_line).account_label
        lines.append(
            BudgetLine(
                account_code=code,
                account_label=label,
                budget_value=budget_line.monthly_values.get(period) if budget_line else None,
                actual_value=actual_line.monthly_values.get(period) if actual_line else None,
            )
        )

    return BudgetComparisonResult(property_code=property_code, period=period, lines=lines)
