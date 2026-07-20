"""Parses an "Equity - *" tab (balance sheet + capital account hybrid) into an
EquityPosition.

Layout observed in the Revolution Labs workbook: column A = GL account code,
column B = indented label, column C = value — but a value is only present on
"Total X" rows (and the per-investor sub-rows immediately above them); every
other row is a blank-value section header. v1 only reads this left block.

A tab also has a second, structurally-inconsistent block further right (a clean
trial balance on the LP tab, a free-form dated ledger on the BHC tab) — that
block is intentionally not parsed; see README "Known workbook quirks".
"""

from __future__ import annotations

import re
from typing import Optional, Tuple

from openpyxl.worksheet.worksheet import Worksheet

from pipeline.models import BalanceSheetLine, EquityPosition

_HEADER_RE = re.compile(r"^(?P<name>.*?)\s*\((?P<code>[^)]+)\)\s*$")


def _split_entity_header(header: str) -> Tuple[str, str]:
    match = _HEADER_RE.match(header.strip())
    if match:
        return match.group("name").strip(), match.group("code").strip()
    return header.strip(), ""


def _lookup_investor_name_above(ws: Worksheet, total_row: int, label_col: int = 2) -> Optional[str]:
    label_above = ws.cell(row=total_row - 1, column=label_col).value
    if isinstance(label_above, str):
        candidate = label_above.strip()
        if candidate and candidate.lower() != "investor":
            return candidate
    return None


def parse_equity_tab(
    ws: Worksheet, label_col: int = 2, value_col: int = 3, max_row: int = 80
) -> EquityPosition:
    header = str(ws.cell(row=1, column=1).value or "")
    entity_name, entity_code = _split_entity_header(header)

    period_raw = str(ws.cell(row=3, column=1).value or "")
    as_of_period = period_raw.split("=", 1)[1].strip() if "=" in period_raw else period_raw.strip()

    balance_sheet = {}
    contributions_by_partner = {}
    distributions_by_partner = {}
    balance_sheet_lines = []

    scan_max = min(ws.max_row, max_row)
    for r in range(1, scan_max + 1):
        raw_label = ws.cell(row=r, column=label_col).value
        value = ws.cell(row=r, column=value_col).value
        if not isinstance(raw_label, str):
            continue
        clean = raw_label.strip()
        if not clean:
            continue

        # Balance-sheet outline: every row down to category/total level, keyed
        # by indentation (raw leading-space count). The deeper "Investor" /
        # per-investor-name breakdown rows (indent 7+) are excluded here —
        # that detail belongs to contributions_by_partner/distributions_by_partner
        # below, not a generic balance sheet line.
        indent = len(raw_label) - len(raw_label.lstrip(" "))
        if indent <= 6:
            code = ws.cell(row=r, column=1).value
            balance_sheet_lines.append(
                BalanceSheetLine(
                    account_code=str(code) if code is not None else "",
                    label=clean,
                    value=float(value) if isinstance(value, (int, float)) else None,
                    indent=indent,
                )
            )

        if value is None or not isinstance(value, (int, float)):
            continue

        upper = clean.upper()
        if upper.startswith("TOTAL "):
            balance_sheet[upper] = float(value)

        lower = clean.lower()
        if "contributions - partner" in lower:
            investor = _lookup_investor_name_above(ws, r, label_col) or clean
            contributions_by_partner[investor] = float(value)
        elif "distributions - partner" in lower:
            investor = _lookup_investor_name_above(ws, r, label_col) or clean
            distributions_by_partner[investor] = float(value)

    return EquityPosition(
        entity_name=entity_name,
        entity_code=entity_code,
        as_of_period=as_of_period,
        balance_sheet=balance_sheet,
        contributions_by_partner=contributions_by_partner,
        distributions_by_partner=distributions_by_partner,
        balance_sheet_lines=balance_sheet_lines,
    )
