"""Deal Dashboard — internal Streamlit app for tracking active/owned CRE deals.

Distinct from the `acquisitions` project (screens incoming OMs pre-acquisition).
"""

import os
import sys
from pathlib import Path
from typing import Optional, Tuple

import streamlit as st

APP_DIR = Path(__file__).parent
if str(APP_DIR) not in sys.path:
    sys.path.insert(0, str(APP_DIR))

from pipeline import source_files
from pipeline.models import BudgetLine, PortfolioSummaryRow
from pipeline.parsers.budget_comparison_report import (
    PTD_ACTUAL_COL,
    PTD_BUDGET_COL,
    YTD_ACTUAL_COL,
    YTD_BUDGET_COL,
    boma_rollup,
    parse_annual_budget_totals,
    parse_budget_comparison_report,
    parse_opex_categories,
    parse_ytd_annual_budget_totals,
    parse_ytd_budget_comparison_report,
)
from pipeline.parsers.cash_accounts import parse_entity_trial_balance, parse_loan_statement
from pipeline.parsers.distribution_workbook import parse_workbook
from pipeline.parsers.rent_roll import parse_rent_roll
from pipeline.property_config import PropertyConfig, discover_properties
from views.branding import render_hero
from views.portfolio_view import render_portfolio
from views.property_detail_view import render_property_detail

DATA_DIR = APP_DIR / "data"

st.set_page_config(page_title="Deal Dashboard", layout="wide", initial_sidebar_state="expanded")


def check_password() -> bool:
    if st.session_state.get("authenticated"):
        return True
    render_hero("Deal Dashboard", "Greatland Realty Partners &mdash; Active Deal Tracking")
    st.write("")
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown("#### Sign In")
        password = st.text_input("Password", type="password", key="password_input")
        if st.button("Sign In", width="stretch"):
            correct = st.secrets.get("APP_PASSWORD") or os.environ.get("APP_PASSWORD", "")
            if password == correct and correct != "":
                st.session_state.authenticated = True
                st.rerun()
            else:
                st.error("Incorrect password.")
    return False


@st.cache_data(show_spinner="Parsing distribution workbook...")
def _cached_parse(path_str: str, mtime: float, property_code: str, data_dir: str):
    cfg = PropertyConfig.load(property_code, data_dir=data_dir)
    return parse_workbook(path_str, cfg)


def _resolve_workbook_path(cfg: PropertyConfig, period: Optional[str]) -> Tuple[Optional[Path], Optional[float]]:
    if not period:
        return None, None
    hint_name = Path(cfg.source_workbook_path_hint).name
    p = source_files.resolve_period_file(cfg, period, hint_name, str(DATA_DIR))
    return (p, p.stat().st_mtime) if p else (None, None)


@st.cache_data(show_spinner="Parsing trial balance...")
def _cached_entity_trial_balance(path_str: str, mtime: float, yardi_codes: tuple):
    return parse_entity_trial_balance(path_str, list(yardi_codes))


def _discover_trial_balance_paths(cfg: PropertyConfig, period: Optional[str]) -> dict:
    """Returns {path_str: mtime} for every trial-balance file resolved for
    this property + period — one per entity (property, venture, co-GP, ...),
    with carry-forward to the nearest earlier period that has any, via
    `source_files.resolve_period_trial_balances`."""
    if not period:
        return {}
    return {str(p): p.stat().st_mtime for p in source_files.resolve_period_trial_balances(cfg, period, str(DATA_DIR))}


@st.cache_data(show_spinner="Parsing loan statement...")
def _cached_loan_statement(path_str: str, mtime: float):
    return parse_loan_statement(path_str)


def _discover_loan_statement_paths(cfg: PropertyConfig, period: Optional[str]) -> dict:
    """Returns {path_str: mtime} for every loan-statement PDF resolved for
    this property + period (with carry-forward to the nearest earlier
    period that has any, via `source_files.resolve_period_loan_statements`)."""
    if not period:
        return {}
    return {str(p): p.stat().st_mtime for p in source_files.resolve_period_loan_statements(cfg, period, str(DATA_DIR))}


@st.cache_data(show_spinner="Parsing rent roll...")
def _cached_rent_roll(path_str: str, mtime: float, property_code: str):
    import openpyxl

    wb = openpyxl.load_workbook(path_str, data_only=True)
    return parse_rent_roll(wb["Report1"] if "Report1" in wb.sheetnames else wb.worksheets[0], property_code)


def _resolve_rent_roll_path(cfg: PropertyConfig, period: Optional[str]) -> Tuple[Optional[Path], Optional[float]]:
    if not period:
        return None, None
    p = source_files.resolve_period_file(cfg, period, "rent_roll.xlsx", str(DATA_DIR))
    return (p, p.stat().st_mtime) if p else (None, None)


@st.cache_data(show_spinner="Parsing budget comparison report...")
def _cached_budget_comparison_report(path_str: str, mtime: float, property_code: str):
    import openpyxl

    wb = openpyxl.load_workbook(path_str, data_only=True)
    ws = wb["Report1"] if "Report1" in wb.sheetnames else wb.worksheets[0]
    return parse_budget_comparison_report(ws, property_code)


@st.cache_data(show_spinner="Parsing budget comparison report...")
def _cached_annual_budget_totals(path_str: str, mtime: float):
    import openpyxl

    wb = openpyxl.load_workbook(path_str, data_only=True)
    ws = wb["Report1"] if "Report1" in wb.sheetnames else wb.worksheets[0]
    return parse_annual_budget_totals(ws)


@st.cache_data(show_spinner="Parsing budget comparison report...")
def _cached_opex_categories(path_str: str, mtime: float, value_col: Optional[int] = None):
    import openpyxl

    wb = openpyxl.load_workbook(path_str, data_only=True)
    ws = wb["Report1"] if "Report1" in wb.sheetnames else wb.worksheets[0]
    return parse_opex_categories(ws) if value_col is None else parse_opex_categories(ws, value_col=value_col)


@st.cache_data(show_spinner="Parsing budget comparison report...")
def _cached_ytd_annual_budget_totals(path_str: str, mtime: float):
    import openpyxl

    wb = openpyxl.load_workbook(path_str, data_only=True)
    ws = wb["Report1"] if "Report1" in wb.sheetnames else wb.worksheets[0]
    return parse_ytd_annual_budget_totals(ws)


@st.cache_data(show_spinner="Parsing budget comparison report...")
def _cached_ytd_budget_comparison_report(path_str: str, mtime: float, property_code: str):
    import openpyxl

    wb = openpyxl.load_workbook(path_str, data_only=True)
    ws = wb["Report1"] if "Report1" in wb.sheetnames else wb.worksheets[0]
    return parse_ytd_budget_comparison_report(ws, property_code)


def _resolve_budget_comparison_path(cfg: PropertyConfig, period: Optional[str]) -> Tuple[Optional[Path], Optional[float]]:
    if not period:
        return None, None
    p = source_files.resolve_period_file(cfg, period, "budget_comparison.xlsx", str(DATA_DIR))
    return (p, p.stat().st_mtime) if p else (None, None)


def _build_portfolio_row(cfg: PropertyConfig, result) -> PortfolioSummaryRow:
    total_cash = None
    total_debt = None
    last_dist_amount = None
    last_dist_as_of = None

    if result:
        cash_values = [pos.total_cash for pos in result.equity.values() if pos.total_cash is not None]
        total_cash = sum(cash_values) if cash_values else None
        total_debt = result.debt.total_outstanding if result.debt else None
        if result.waterfall:
            top_tier_cfg = cfg.top_level_tier()
            top = result.waterfall.tiers.get(top_tier_cfg.tier_id) if top_tier_cfg else None
            if top is None and result.waterfall.tiers:
                top = next(iter(result.waterfall.tiers.values()))
            if top:
                last_dist_amount = top.distribution_recommendation
                last_dist_as_of = top.as_of_label

    return PortfolioSummaryRow(
        property_code=cfg.property_code,
        display_name=cfg.display(),
        address=cfg.property_address,
        market=cfg.market,
        investor_names=cfg.investor_display_names(),
        total_cash=total_cash,
        total_debt_outstanding=total_debt,
        last_distribution_amount=last_dist_amount,
        last_distribution_as_of=last_dist_as_of,
    )


def _process_uploads(cfg: PropertyConfig, uploaded_files) -> Tuple[Optional[str], list]:
    """Classifies + saves every file in a batch upload; returns (newest
    period touched, [(level, message), ...]). The messages are returned
    rather than written directly via st.sidebar.warning/success because the
    caller immediately triggers a st.rerun() to refresh the period picker —
    messages emitted before that rerun are wiped before a user can read them,
    so they need to be stashed in session_state and rendered on the next
    run instead."""
    newest_period = None
    messages = []
    for f in uploaded_files:
        data = f.getvalue()
        classified = source_files.classify_upload(data, f.name, cfg)
        if classified.file_type == "unknown" or not classified.period:
            messages.append(("warning", classified.error or f"Couldn't process {f.name}."))
            continue
        source_files.save_classified_upload(classified, data, cfg, str(DATA_DIR))
        messages.append(("success", f"{f.name} → {classified.file_type.replace('_', ' ')} ({classified.period})"))
        if newest_period is None or classified.period > newest_period:
            newest_period = classified.period
    return newest_period, messages


_TEMPLATE_SHEET_MAP = {
    "cash_flow": "Cash Flow",
    "equity_lp": "Equity - LP",
    "equity_bhc": "Equity - BHC",
    "debt": "Loan Interest",
    "waterfall_property": "Distribution Recommendation",
    "waterfall_cogp": "Distribution Recommendation BHC",
    "budget_current": "Budget",
}
_ROLE_OPTIONS = ["LP", "GP", "Sponsor", "co-GP", "co-GP member"]


def _render_how_to_use() -> None:
    st.markdown(
        """
**Views** (the tabs above)
- **Portfolio** — every active property, one row each, with headline cash/debt/distribution figures.
- **Property Detail** — everything for the **Active Property** picked above. Pick a month/quarter from **Viewing Period** next to it.
- **Properties** — add a new property or edit an existing one's basics. See below.

**Uploading files**
Drop files into **Update Source Files** (sidebar) — each one is auto-detected by its content (not its filename), so you don't need to sort them first. Drop as many as you want at once. Recognized types: distribution workbook, trial balance (one file per entity — e.g. property, venture, co-GP), rent roll, Yardi Budget Comparison report, and loan servicer statement PDFs. A file is filed under whatever period it's dated as of — if a type isn't re-uploaded in a given month, the app shows the last known version instead of going blank.

If a file gets rejected, the message right above the uploader explains why (usually "couldn't find an as-of date" or "doesn't look like any known file type") — it stays on screen until the next upload.

**Removing a file** — open **Reset Source Files** (sidebar), pick the period, and either remove one file or clear the whole period. Each action needs you to type the exact file/period name first — deletes can't be undone from here.

**Sections** (the pills below the property header, inside Property Detail)
- **Summary** — the one-glance view: cash, NOI, debt, occupancy, with jump links into the detail sections.
- **Cash** — every cash/escrow/reserve account, by source.
- **Equity & Capital** — contributions, distributions, and capital balances per entity.
- **Balance Sheet** — full balance sheet per entity, from the distribution workbook's equity tabs.
- **Rent Roll** — occupancy, lease terms, current rent/CAM, and upcoming rent steps per suite.
- **Debt & Loans** — forecast (distribution workbook) vs. actual (loan statements) balances and rates, side by side.
- **Distribution Waterfall** — the JV distribution calc per ownership tier.
- **Budget vs. Actuals** — two toggles: Period-to-Date (the selected month) vs. Year-to-Date, and Summary (P&L rollup, plus a BOMA-category OpEx breakdown) vs. Detailed (account-level). Year-to-Date also shows a Reforecast column (the distribution workbook's own current-year figure) alongside Kardin's once-per-year Annual Budget.
- **Hold/Sell Assumptions** — see below.
- **Sources & Uses** / **Leasing & Investment Outlook** — placeholders until the leasing/investment model is finalized.
- **Reconciliation** — cross-checks each uploaded trial balance's own figures (AR/AP/mortgage/cash) against the other source files that should describe the same real-world numbers; flags anything off by more than a small tolerance (could just be a timing difference between files).

**Hold/Sell Assumptions**
This section holds the inputs for the property's hold/sell model (inflation, vacancy, market leasing terms, cap rate, hold period, refinance) so anyone can view or change them without opening Excel. **Save Assumptions** persists them for next time. **Download Excel Model** produces a fresh workbook with those assumptions plus the latest real data (rent roll, actuals, debt, equity) already filled in — all the actual math (rollover schedule, debt amortization, IRR) lives in that workbook's own formulas, not in the app, so the exported file is always the authoritative calculation.

**Adding a property**
Use the **Properties** tab: fill in the basics, upload a hero photo, Yardi entity codes, and (optionally) a single ownership tier with its investors, then **Save Property**. If GitHub isn't connected (see the banner at the top of that tab), everything only saves to this machine's local disk — download the YAML and commit it (plus the photo) yourself, or ask to have `[github]` secrets set up so saves auto-deploy.
"""
    )


def _render_properties_tab(properties: list) -> None:
    from pipeline import property_writer

    st.caption(
        "Add a new property, or edit an existing one's basics — no YAML editing "
        "required. Loan/JV abstracts still get added separately, once they exist."
    )

    if property_writer.github_configured():
        st.success("GitHub connected — saved configs deploy automatically in ~2 min.", icon="✅")
    else:
        st.warning(
            "GitHub not connected — configs save locally only, which won't persist on the "
            "hosted app past the next redeploy. Add a `[github]` token + repo to this app's "
            "Streamlit secrets to enable auto-deploy (same as ga-automation does), or use the "
            "download button below and commit the file yourself.",
            icon="⚠️",
        )

    codes = [c.property_code for c in properties]
    names = {c.property_code: c.display() for c in properties}
    edit_choice = st.selectbox(
        "Edit existing or create new",
        options=["+ Create new property"] + codes,
        format_func=lambda c: c if c == "+ Create new property" else f"{names[c]}  ({c})",
        key="prop_setup_edit_select",
    )
    is_new = edit_choice == "+ Create new property"
    edit_cfg = None if is_new else next((p for p in properties if p.property_code == edit_choice), None)
    # Every widget key below is suffixed with `edit_choice` — switching the edit
    # target must force brand-new widgets, since Streamlit ignores a freshly
    # computed `value=` on any widget whose key already has session state from
    # a previous render (confirmed gotcha, see ga-automation's own CLAUDE.md).
    ek = edit_choice

    def ef(field, default=""):
        if edit_cfg is None:
            return default
        return getattr(edit_cfg, field, default) or default

    st.markdown("#### Basics")
    col1, col2 = st.columns(2)
    with col1:
        property_code = st.text_input(
            "Property Code (folder name — lowercase, no spaces)",
            value="" if is_new else edit_cfg.property_code,
            disabled=not is_new,
            key=f"prop_code_{ek}",
            help=None if is_new else "Can't be changed after creation — it's the data/<code>/ folder name.",
        )
        property_name = st.text_input("Legal Owning Entity", value=ef("property_name"), key=f"prop_name_{ek}")
        property_display_name = st.text_input("Display Name (shown in the UI)", value=ef("property_display_name"), key=f"prop_display_{ek}")
        property_address = st.text_input("Address", value=ef("property_address"), key=f"prop_address_{ek}")
    with col2:
        property_type = st.text_input("Property Type (e.g. Office, Life Science)", value=ef("property_type"), key=f"prop_type_{ek}")
        market = st.text_input("Market", value=ef("market"), key=f"prop_market_{ek}")
        submarket = st.text_input("Submarket", value=ef("submarket"), key=f"prop_submarket_{ek}")
        state = st.text_input("State", value=ef("state"), key=f"prop_state_{ek}")
    management_company = st.text_input(
        "Management Company", value=ef("management_company", "Greatland Realty Partners"), key=f"prop_mgmt_{ek}"
    )
    active = st.checkbox("Active", value=True if edit_cfg is None else edit_cfg.active, key=f"prop_active_{ek}")

    st.markdown("#### Property Photo")
    st.caption("Shown in the hero banner at the top of every page for this property. JPG, PNG, or WebP.")
    assets_dir = str(APP_DIR / "assets")
    existing_photo_path = None
    if edit_cfg:
        for ext in (".jpg", ".jpeg", ".png", ".webp"):
            candidate = Path(assets_dir) / f"{edit_cfg.property_code}_hero{ext}"
            if candidate.exists():
                existing_photo_path = candidate
                break
    if existing_photo_path:
        st.image(str(existing_photo_path), caption="Current photo", width=280)
    photo_upload = st.file_uploader(
        "Upload a new photo" if existing_photo_path else "Upload a photo",
        type=["jpg", "jpeg", "png", "webp"],
        key=f"prop_photo_{ek}",
    )

    st.markdown("#### Yardi Entity Codes")
    st.caption(
        "The real Yardi export codes for this property's entities (property/venture/co-GP) — "
        "used to filter multi-property source files down to this one. Comma-separated; often "
        "differs from the Property Code above (confirmed for Revolution Labs: dashboard code "
        "\"revlabspm\" vs. real Yardi codes \"revlabpm\"/\"revlabvn\"/\"bh1050jv\"/\"revlabs\")."
    )
    default_yardi = ", ".join(edit_cfg.yardi_codes) if edit_cfg else ""
    yardi_codes_raw = st.text_input("Yardi Codes", value=default_yardi, key=f"prop_yardi_{ek}")
    yardi_codes = [c.strip() for c in yardi_codes_raw.split(",") if c.strip()]

    with st.expander("Distribution Workbook Tab Names (advanced)", expanded=False):
        st.caption(
            "Tab names inside the quarterly distribution workbook — only change these if this "
            "property's workbook uses different tab names than the ones below."
        )
        existing_map = edit_cfg.sheet_map if edit_cfg and edit_cfg.sheet_map else {}
        sheet_map = {}
        for key, default_label in _TEMPLATE_SHEET_MAP.items():
            sheet_map[key] = st.text_input(
                key.replace("_", " ").title(), value=existing_map.get(key, default_label), key=f"prop_sheetmap_{key}_{ek}"
            )

    st.markdown("#### Ownership / JV Structure")
    st.caption(
        "One tier is enough for a simple LP/GP split. Multi-tier (co-GP) structures — like "
        "Revolution Labs' own LP/GP + Co-GP split — aren't supported by this form yet; save "
        "this as a single tier below, then edit config.yaml directly to add a nested tier."
    )
    top_tier = edit_cfg.top_level_tier() if edit_cfg else None
    tier_label = st.text_input(
        "Tier Label (shown as the Distribution Waterfall tab name)",
        value=top_tier.label() if top_tier else "LP/GP",
        key=f"prop_tier_label_{ek}",
    )
    tier_entity = st.text_input(
        "Distributing Entity Name", value=top_tier.distributing_entity if top_tier else "", key=f"prop_tier_entity_{ek}"
    )

    if st.session_state.get("prop_investor_rows_for") != ek:
        st.session_state.prop_investor_rows = [
            {
                "display_name": inv.display_name,
                "legal_entity": inv.legal_entity,
                "ownership_pct": inv.ownership_pct * 100,
                "role": inv.role,
            }
            for inv in (top_tier.investors if top_tier else [])
        ] or [{"display_name": "", "legal_entity": "", "ownership_pct": 0.0, "role": "LP"}]
        st.session_state.prop_investor_rows_for = ek

    st.markdown("**Investors**")
    hdr1, hdr2, hdr3, hdr4, _hdr5 = st.columns([3, 3, 2, 2, 1])
    hdr1.caption("Investor Name")
    hdr2.caption("Legal Entity")
    hdr3.caption("Ownership %")
    hdr4.caption("Role")

    row_to_remove = None
    for i, row in enumerate(st.session_state.prop_investor_rows):
        c1, c2, c3, c4, c5 = st.columns([3, 3, 2, 2, 1])
        row["display_name"] = c1.text_input("Investor Name", value=row["display_name"], key=f"inv_name_{ek}_{i}", label_visibility="collapsed")
        row["legal_entity"] = c2.text_input("Legal Entity", value=row["legal_entity"], key=f"inv_entity_{ek}_{i}", label_visibility="collapsed")
        row["ownership_pct"] = c3.number_input("Ownership %", value=row["ownership_pct"], step=1.0, key=f"inv_pct_{ek}_{i}", label_visibility="collapsed")
        role_idx = _ROLE_OPTIONS.index(row["role"]) if row["role"] in _ROLE_OPTIONS else 0
        row["role"] = c4.selectbox("Role", _ROLE_OPTIONS, index=role_idx, key=f"inv_role_{ek}_{i}", label_visibility="collapsed")
        if c5.button("✕", key=f"inv_remove_{ek}_{i}"):
            row_to_remove = i
    if row_to_remove is not None:
        st.session_state.prop_investor_rows.pop(row_to_remove)
        st.rerun()
    if st.button("+ Add investor", key=f"prop_add_investor_{ek}"):
        st.session_state.prop_investor_rows.append({"display_name": "", "legal_entity": "", "ownership_pct": 0.0, "role": "LP"})
        st.rerun()

    st.divider()
    if st.button("Save Property", type="primary", key=f"prop_save_{ek}"):
        code_clean = property_code.strip()
        if not code_clean:
            st.error("Property Code is required.")
        elif is_new and code_clean in codes:
            st.error(f'A property with code "{code_clean}" already exists.')
        else:
            investor_dicts = [
                {
                    "display_name": r["display_name"],
                    "legal_entity": r["legal_entity"],
                    "ownership_pct": r["ownership_pct"] / 100,
                    "role": r["role"],
                    "sub_tier": None,
                }
                for r in st.session_state.prop_investor_rows
                if r["display_name"].strip()
            ]
            ownership_tiers = (
                [
                    {
                        "tier_id": "property",
                        "distributing_entity": tier_entity,
                        "parent_tier": None,
                        "display_label": tier_label,
                        "investors": investor_dicts,
                    }
                ]
                if investor_dicts
                else []
            )
            # Loans/JV documents have no form fields (abstracts are a separate,
            # later, one-time step) — carry forward whatever's already there
            # when editing, rather than silently wiping them out on save.
            loans = (
                [{"tranche_name": l.tranche_name, "lender": l.lender, "abstract_file": l.abstract_file} for l in edit_cfg.loans]
                if edit_cfg
                else []
            )
            jv_documents = (
                [{"name": d.name, "abstract_file": d.abstract_file} for d in edit_cfg.jv_documents] if edit_cfg else []
            )

            config_dict = property_writer.build_config_dict(
                property_code=code_clean,
                property_name=property_name,
                property_display_name=property_display_name,
                property_address=property_address,
                property_type=property_type,
                market=market,
                submarket=submarket,
                state=state,
                management_company=management_company,
                active=active,
                sheet_map=sheet_map,
                yardi_codes=yardi_codes,
                ownership_tiers=ownership_tiers,
                loans=loans,
                jv_documents=jv_documents,
            )
            yaml_content = property_writer.config_to_yaml(config_dict)

            local_ok, local_msg = property_writer.save_local(code_clean, yaml_content, str(DATA_DIR))
            if local_ok:
                st.success(f"Saved locally: {local_msg}")
            else:
                st.error(f"Local save failed: {local_msg}")

            if property_writer.github_configured():
                gh_ok, gh_msg = property_writer.save_to_github(code_clean, yaml_content)
                (st.success if gh_ok else st.error)(gh_msg)

            if photo_upload is not None:
                photo_bytes = photo_upload.getvalue()
                photo_ext = "." + photo_upload.name.rsplit(".", 1)[-1].lower()
                photo_local_ok, photo_local_msg = property_writer.save_hero_photo_local(
                    code_clean, photo_bytes, photo_ext, assets_dir
                )
                if photo_local_ok:
                    st.success(f"Photo saved locally: {photo_local_msg}")
                else:
                    st.error(f"Photo local save failed: {photo_local_msg}")
                if property_writer.github_configured():
                    photo_gh_ok, photo_gh_msg = property_writer.save_hero_photo_to_github(code_clean, photo_bytes, photo_ext)
                    (st.success if photo_gh_ok else st.error)(photo_gh_msg)

            st.download_button(
                "Download config.yaml",
                data=yaml_content,
                file_name=f"{code_clean}_config.yaml",
                mime="text/yaml",
                key=f"prop_download_{ek}",
            )
            if local_ok:
                st.info("Switch tabs (or reload) to see it in the Active Property list.")


def main() -> None:
    for key, default in {
        "selected_property": None,
        "upload_epoch": {},
        "upload_messages": {},
    }.items():
        if key not in st.session_state:
            st.session_state[key] = default

    properties = discover_properties(str(DATA_DIR))

    st.sidebar.markdown(
        "<h2 style='color:#1A5C22;'>Deal Dashboard</h2>", unsafe_allow_html=True
    )

    if not properties:
        render_hero("Deal Dashboard", "Greatland Realty Partners &mdash; Active Deal Tracking")
        st.info("No properties configured yet — add your first one in the **Properties** tab below.")
        tab_howto_empty, tab_properties_empty = st.tabs(["How to Use", "Properties"])
        with tab_howto_empty:
            _render_how_to_use()
        with tab_properties_empty:
            _render_properties_tab(properties)
        return

    codes = [c.property_code for c in properties]
    names = {c.property_code: c.display() for c in properties}

    # Active property persists per-browser via the URL's ?property= query param —
    # so the app reopens on whatever each person was last looking at, and a
    # property is linkable/bookmarkable. Falls back to the first property if the
    # URL has nothing (or something stale, e.g. a deleted property).
    if "active_property_code" not in st.session_state:
        qp_code = st.query_params.get("property")
        st.session_state.active_property_code = qp_code if qp_code in codes else codes[0]
    if st.session_state.active_property_code not in codes:
        st.session_state.active_property_code = codes[0]

    cfg = next(c for c in properties if c.property_code == st.session_state.active_property_code)
    badges = [b for b in [cfg.market, cfg.property_type] if b]
    render_hero(cfg.display(), cfg.property_address, badges, photo_code=cfg.property_code)

    sel_col, period_col, _spacer_col = st.columns([2, 2, 4])
    with sel_col:
        selected_code = st.selectbox(
            "Active Property",
            codes,
            index=codes.index(st.session_state.active_property_code),
            format_func=lambda c: names[c],
            key="active_property_selectbox",
        )
        if selected_code != st.session_state.active_property_code:
            st.session_state.active_property_code = selected_code
            st.query_params["property"] = selected_code
            st.rerun()
    if st.query_params.get("property") != st.session_state.active_property_code:
        st.query_params["property"] = st.session_state.active_property_code

    selected_code = st.session_state.active_property_code
    cfg = next(c for c in properties if c.property_code == selected_code)

    period_key = f"period_select_{selected_code}"
    if "pending_period" in st.session_state:
        st.session_state[period_key] = st.session_state.pop("pending_period")

    periods = source_files.list_periods(cfg, str(DATA_DIR))
    selected_period = None
    with period_col:
        if periods:
            if period_key not in st.session_state or st.session_state[period_key] not in periods:
                st.session_state[period_key] = periods[0]
            selected_period = st.selectbox("Viewing Period", periods, key=period_key)
        else:
            st.selectbox("Viewing Period", ["No periods yet"], disabled=True)

    st.sidebar.markdown("---")
    with st.sidebar.expander("Update Source Files"):
        st.caption(
            "Drop in an updated distribution workbook, trial balance, rent roll, "
            "Yardi Budget Comparison report, or loan statement PDF(s) — each file "
            "is auto-detected by its content and filed under the period it's "
            "dated as of."
        )
        for level, message in st.session_state.upload_messages.get(selected_code, []):
            getattr(st, level)(message)

        epoch = st.session_state.upload_epoch.get(selected_code, 0)
        uploaded_files = st.file_uploader(
            "Files",
            type=["xlsx", "pdf"],
            accept_multiple_files=True,
            key=f"multi_upload_{selected_code}_{epoch}",
            label_visibility="collapsed",
        )
        if uploaded_files:
            newest_period, messages = _process_uploads(cfg, uploaded_files)
            st.session_state.upload_messages[selected_code] = messages
            if newest_period:
                st.session_state.pending_period = newest_period
            st.session_state.upload_epoch[selected_code] = epoch + 1
            st.rerun()

    with st.sidebar.expander("Reset Source Files"):
        st.caption(
            "Remove a file that was uploaded by mistake, or clear an entire period. "
            "Deleting can't be undone from here, so each action needs the exact "
            "file or period name typed in before it activates."
        )
        if not periods:
            st.caption("No periods yet.")
        else:
            reset_period_key = f"reset_period_select_{selected_code}"
            if reset_period_key not in st.session_state or st.session_state[reset_period_key] not in periods:
                st.session_state[reset_period_key] = selected_period or periods[0]
            reset_period = st.selectbox("Period", periods, key=reset_period_key)

            reset_files = source_files.list_period_contents(cfg, reset_period, str(DATA_DIR))
            if not reset_files:
                st.caption(f"No files for {reset_period}.")
            else:
                for f in reset_files:
                    col_label, col_confirm, col_button = st.columns([3, 2, 1])
                    col_label.markdown(f"**{source_files.describe_period_file(f)}**\n\n{f.name}")
                    confirm_key = f"reset_confirm_{selected_code}_{reset_period}_{f.name}"
                    typed = col_confirm.text_input(
                        "Confirm", key=confirm_key, placeholder=f.name, label_visibility="collapsed"
                    )
                    if col_button.button("Remove", key=f"reset_btn_{confirm_key}", disabled=(typed != f.name)):
                        source_files.delete_period_file(f)
                        st.rerun()

                st.divider()
                st.caption(f"Clear every file for {reset_period}:")
                col_confirm, col_button = st.columns([3, 1])
                clear_key = f"clear_period_confirm_{selected_code}_{reset_period}"
                typed_period = col_confirm.text_input(
                    "Confirm period",
                    key=clear_key,
                    placeholder=f"Type {reset_period} to confirm",
                    label_visibility="collapsed",
                )
                if col_button.button(
                    "Delete this period", key=f"clear_period_btn_{clear_key}", disabled=(typed_period != reset_period)
                ):
                    source_files.delete_period(cfg, reset_period, str(DATA_DIR))
                    st.rerun()

    tab_howto, tab_portfolio, tab_detail, tab_properties = st.tabs(
        ["How to Use", "Portfolio", "Property Detail", "Properties"]
    )

    with tab_howto:
        _render_how_to_use()

    with tab_portfolio:
        rows = []
        for p_cfg in properties:
            p_periods = source_files.list_periods(p_cfg, str(DATA_DIR))
            latest_period = p_periods[0] if p_periods else None
            p_path, p_mtime = _resolve_workbook_path(p_cfg, latest_period)
            p_result = _cached_parse(str(p_path), p_mtime, p_cfg.property_code, str(DATA_DIR)) if p_path else None
            rows.append(_build_portfolio_row(p_cfg, p_result))
        render_portfolio(rows)

    with tab_detail:
        if not periods:
            st.info(
                f"No source files yet for **{cfg.display()}**. Use “Update Source Files” "
                "in the sidebar to upload the distribution workbook, trial balance, rent roll, "
                "and/or loan statements."
            )
        else:
            path, mtime = _resolve_workbook_path(cfg, selected_period)
            result = None
            if path is None:
                st.info(f"No distribution workbook found for **{cfg.display()}** as of {selected_period}.")
            else:
                try:
                    result = _cached_parse(str(path), mtime, cfg.property_code, str(DATA_DIR))
                except Exception as exc:
                    st.error(f"Failed to parse workbook: {exc}")

            opex_categories: dict = {}
            boma_opex: dict = {}
            if result is not None:
                bc_path, bc_mtime = _resolve_budget_comparison_path(cfg, selected_period)
                if bc_path is not None:
                    try:
                        bc_result = _cached_budget_comparison_report(str(bc_path), bc_mtime, cfg.property_code)
                        if bc_result.lines:
                            result.budget_comparison = bc_result

                        annual_totals = _cached_annual_budget_totals(str(bc_path), bc_mtime)
                        if len(annual_totals) == 8 and result.annual_budget_summary:
                            result.annual_budget_summary.lines = [
                                BudgetLine(
                                    account_code=line.account_code,
                                    account_label=line.account_label,
                                    # Kardin's static once-per-year budget replaces budget_value; the
                                    # distribution workbook's own figure it displaces is kept as the
                                    # live reforecast, not discarded.
                                    budget_value=annual_totals.get(line.account_code, line.budget_value),
                                    actual_value=line.actual_value,
                                    reforecast_value=line.budget_value,
                                )
                                for line in result.annual_budget_summary.lines
                            ]

                        ytd_detail = _cached_ytd_budget_comparison_report(str(bc_path), bc_mtime, cfg.property_code)
                        if ytd_detail.lines:
                            result.annual_budget_detail = ytd_detail

                        opex_categories = _cached_opex_categories(str(bc_path), bc_mtime)
                        boma_opex = {
                            "ptd": {
                                "actual": boma_rollup(_cached_opex_categories(str(bc_path), bc_mtime, PTD_ACTUAL_COL)),
                                "budget": boma_rollup(_cached_opex_categories(str(bc_path), bc_mtime, PTD_BUDGET_COL)),
                            },
                            "ytd": {
                                "actual": boma_rollup(_cached_opex_categories(str(bc_path), bc_mtime, YTD_ACTUAL_COL)),
                                "budget": boma_rollup(_cached_opex_categories(str(bc_path), bc_mtime, YTD_BUDGET_COL)),
                            },
                        }
                    except Exception as exc:
                        st.error(f"Failed to parse budget comparison report ({Path(bc_path).name}): {exc}")

            entity_trial_balances = []
            for path_str, tb_mtime in _discover_trial_balance_paths(cfg, selected_period).items():
                try:
                    entity_trial_balances.extend(_cached_entity_trial_balance(path_str, tb_mtime, tuple(cfg.yardi_codes)))
                except Exception as exc:
                    st.error(f"Failed to parse trial balance ({Path(path_str).name}): {exc}")

            cash_accounts = [acct for entity in entity_trial_balances for acct in entity.cash_accounts]

            loan_statements = []
            for path_str, ls_mtime in _discover_loan_statement_paths(cfg, selected_period).items():
                try:
                    stmt = _cached_loan_statement(path_str, ls_mtime)
                    if stmt:
                        loan_statements.append(stmt)
                except Exception as exc:
                    st.error(f"Failed to parse loan statement ({Path(path_str).name}): {exc}")

            if not cash_accounts and loan_statements:
                from pipeline.parsers.cash_accounts import loan_statement_cash_accounts

                for stmt in loan_statements:
                    cash_accounts.extend(loan_statement_cash_accounts(stmt))

            rr_path, rr_mtime = _resolve_rent_roll_path(cfg, selected_period)
            rent_roll = None
            if rr_path is not None:
                try:
                    rent_roll = _cached_rent_roll(str(rr_path), rr_mtime, cfg.property_code)
                except Exception as exc:
                    st.error(f"Failed to parse rent roll: {exc}")

            render_property_detail(
                cfg, result, cash_accounts, rent_roll, loan_statements, entity_trial_balances, str(DATA_DIR),
                opex_categories, boma_opex,
            )

    with tab_properties:
        _render_properties_tab(properties)


if not check_password():
    st.stop()

main()
